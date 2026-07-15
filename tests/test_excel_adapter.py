from datetime import date, time
from shutil import copyfile
from types import SimpleNamespace

import numpy as np
import pytest
from openpyxl import Workbook

from geoworkbench.data.excel_adapter import (
    ExcelImportError,
    ExcelImportPlan,
    import_excel,
    probe_excel,
)


def make_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Logging"
    sheet.append(["report", None, None])
    sheet.append(["DEPTH [m]", "C1 [%]", "ROP [m/h]"])
    sheet.append([100.0, 1.2, 10.0])
    sheet.append([100.5, 1.4, 11.0])
    workbook.create_sheet("Empty")
    workbook.save(path)


def test_probe_and_import_excel_sheet_with_header_row(tmp_path) -> None:
    source = tmp_path / "logging.xlsx"
    make_workbook(source)

    probe = probe_excel(source, sheet_name="Logging", header_row=2)
    result = import_excel(
        source,
        ExcelImportPlan("Logging", header_row=2, index_column="DEPTH [m]"),
    )

    assert probe.sheet_names == ("Logging", "Empty")
    assert probe.columns == ("DEPTH [m]", "C1 [%]", "ROP [m/h]")
    assert result.dataset.source_path == source
    assert result.dataset.name == "logging — Logging"
    np.testing.assert_allclose(result.dataset.depth, [100.0, 100.5])
    assert result.dataset.curve_by_mnemonic("C1").metadata.unit == "%"  # type: ignore[union-attr]


def test_import_excel_combines_native_date_and_time_cells(tmp_path) -> None:
    source = tmp_path / "time.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Time"
    sheet.append(["DATE", "TIME", "C1"])
    sheet.append([date(2026, 7, 15), time(10, 0), 1.0])
    workbook.save(source)

    result = import_excel(
        source,
        ExcelImportPlan(
            "Time",
            index_column="DATE",
            time_column="TIME",
            timezone="Asia/Oral",
        ),
    )

    np.testing.assert_array_equal(
        result.dataset.active_index.values,
        np.array(["2026-07-15T05:00:00"], dtype="datetime64[ns]"),
    )


def test_excel_rejects_missing_sheet_and_invalid_extension(tmp_path) -> None:
    source = tmp_path / "logging.xlsx"
    make_workbook(source)

    with pytest.raises(ExcelImportError, match="не найден"):
        probe_excel(source, sheet_name="Unknown")
    text = tmp_path / "logging.ods"
    text.write_text("unsupported", encoding="utf-8")
    with pytest.raises(ExcelImportError, match="XLS/XLSX/XLSM"):
        probe_excel(text)


def test_xls_uses_isolated_libreoffice_conversion(tmp_path, monkeypatch) -> None:
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"legacy binary workbook")
    converted_template = tmp_path / "converted.xlsx"
    make_workbook(converted_template)
    commands: list[list[str]] = []

    monkeypatch.setattr(
        "geoworkbench.data.excel_adapter._find_libreoffice", lambda: "/usr/bin/soffice"
    )

    def convert(command, **kwargs):
        commands.append(command)
        output = command[command.index("--outdir") + 1]
        copyfile(converted_template, f"{output}/legacy.xlsx")
        return SimpleNamespace(returncode=0, stdout="", stderr="")

    monkeypatch.setattr("geoworkbench.data.excel_adapter.subprocess.run", convert)

    result = import_excel(source, ExcelImportPlan("Logging", 2, "DEPTH [m]"))

    np.testing.assert_allclose(result.dataset.depth, [100.0, 100.5])
    assert "--headless" in commands[0]
    assert any(item.startswith("-env:UserInstallation=") for item in commands[0])


def test_formula_workbook_requires_libreoffice_for_recalculation(tmp_path, monkeypatch) -> None:
    source = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["DEPTH", "TOTAL"])
    sheet.append([100.0, "=1+2"])
    workbook.save(source)
    monkeypatch.setattr("geoworkbench.data.excel_adapter._find_libreoffice", lambda: None)

    with pytest.raises(ExcelImportError, match="libreoffice-calc"):
        import_excel(source, ExcelImportPlan(sheet.title, 1, "DEPTH"))
