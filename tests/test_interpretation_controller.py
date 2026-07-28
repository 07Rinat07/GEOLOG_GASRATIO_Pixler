import csv
import json

import numpy as np
import pytest
from openpyxl import load_workbook

from geoworkbench.data.interpretation_export import (
    export_interpretation_csv,
    export_interpretation_excel,
)
from geoworkbench.domain.models import (
    Dataset,
    DatasetKind,
    DepthDomain,
    InterpretationInterval,
    Well,
    WellInterpretation,
)
from geoworkbench.project.interpretation_controller import InterpretationController
from geoworkbench.project.session import ProjectSession


def make_controller() -> InterpretationController:
    session = ProjectSession()
    session.add_dataset(
        Dataset(
            "dataset",
            "Well A",
            DatasetKind.GTI,
            DepthDomain.MD,
            np.array([100.0, 200.0]),
        )
    )
    session.dirty = False
    return InterpretationController(session)


def test_interpretation_controller_crud_and_history() -> None:
    controller = make_controller()
    interpretation = controller.add_interpretation(
        "Primary",
        description="Main geological interpretation",
    )
    first = controller.add_interval(
        100.0,
        150.0,
        "Reservoir",
        "Sand A",
        color="#facc15",
        comment="Good gas response",
    )
    second = controller.add_interval(
        150.0,
        200.0,
        "Reservoir",
        "Sand B",
    )

    assert controller.current_interpretation() is interpretation
    assert controller.available_intervals() == (first, second)
    assert controller.can_undo is True

    controller.update_interval(
        first.interval_id,
        top_depth=100.0,
        bottom_depth=140.0,
        interval_type="Reservoir",
        label="Sand A1",
        color="#facc15",
        comment="Updated",
    )
    assert first.bottom_depth == 140.0
    assert first.label == "Sand A1"

    controller.undo()
    restored = controller.available_intervals()[0]
    assert restored.bottom_depth == 150.0
    assert restored.label == "Sand A"

    controller.redo()
    redone = controller.available_intervals()[0]
    assert redone.bottom_depth == 140.0
    assert redone.label == "Sand A1"
    assert controller.session.dirty is True


def test_interpretation_controller_rejects_invalid_and_same_type_overlap() -> None:
    controller = make_controller()
    controller.add_interpretation("Primary")
    controller.add_interval(110.0, 150.0, "Reservoir", "A")

    with pytest.raises(ValueError, match="пересекается"):
        controller.add_interval(140.0, 160.0, "Reservoir", "B")

    # A different semantic type may intentionally overlay the same depth interval.
    controller.add_interval(140.0, 160.0, "Risk", "Uncertain boundary")

    with pytest.raises(ValueError, match="диапазон"):
        controller.add_interval(90.0, 110.0, "Reservoir", "Outside")
    with pytest.raises(ValueError, match="#RRGGBB"):
        controller.add_interval(160.0, 170.0, "Reservoir", "Bad color", color="red")


def test_interpretation_exports_json_csv_and_excel(tmp_path) -> None:
    controller = make_controller()
    controller.add_interpretation("Primary", description="Export test")
    controller.add_interval(
        100.0,
        120.0,
        "Fluid",
        "Gas-bearing",
        color="#60a5fa",
        comment="High confidence",
    )

    json_path = controller.export_current(tmp_path / "intervals", "json")
    csv_path = controller.export_current(tmp_path / "intervals", "csv")
    xlsx_path = controller.export_current(tmp_path / "intervals", "xlsx")

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["schema"] == "geolog.interpretation.intervals.v1"
    assert payload["well_name"] == "Well A"
    assert payload["interpretation"]["intervals"][0]["label"] == "Gas-bearing"

    csv_text = csv_path.read_text(encoding="utf-8-sig")
    assert "Gas-bearing" in csv_text
    assert "High confidence" in csv_text

    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook["Intervals"]["D2"].value == "Fluid"
    assert workbook["Metadata"]["B3"].value == "Primary"


def test_interpretation_csv_and_xlsx_protect_formula_like_text_only(tmp_path) -> None:
    interpretation = WellInterpretation(
        interpretation_id="\t=INTERPRETATION_ID",
        name="\ufeff@NAME",
        description="\n-DESCRIPTION",
        intervals=[
            InterpretationInterval(
                interval_id=" \t=INTERVAL_ID",
                top_depth=-5.0,
                bottom_depth=10.0,
                interval_type="\n+TYPE",
                label="\ufeff@LABEL",
                color="#aabbcc",
                comment="\n-COMMENT",
            )
        ],
    )
    csv_path = export_interpretation_csv(
        interpretation,
        tmp_path / "formula-guard.csv",
    )
    xlsx_path = export_interpretation_excel(
        interpretation,
        tmp_path / "formula-guard.xlsx",
        well_name="=WELL",
    )

    with csv_path.open(encoding="utf-8-sig", newline="") as stream:
        row = next(csv.DictReader(stream))
    assert row["interval_id"] == "' \t=INTERVAL_ID"
    assert row["top_depth"] == "-5.0"
    assert row["interval_type"] == "'\n+TYPE"
    assert row["label"] == "'\ufeff@LABEL"
    assert row["color"] == "#aabbcc"
    assert row["comment"] == "'\n-COMMENT"

    workbook = load_workbook(xlsx_path, data_only=False)
    intervals = workbook["Intervals"]
    metadata = workbook["Metadata"]
    assert intervals["A2"].value == "' \t=INTERVAL_ID"
    assert intervals["B2"].value == -5.0
    assert intervals["B2"].data_type == "n"
    assert intervals["D2"].value == "'\n+TYPE"
    assert intervals["E2"].value == "'\ufeff@LABEL"
    assert intervals["F2"].value == "#aabbcc"
    assert intervals["G2"].value == "'\n-COMMENT"
    assert metadata["B1"].value == "'=WELL"
    assert metadata["B2"].value == "'\t=INTERPRETATION_ID"
    assert metadata["B3"].value == "'\ufeff@NAME"
    assert metadata["B4"].value == "'\n-DESCRIPTION"
    assert all(
        cell.data_type != "f"
        for sheet in workbook.worksheets
        for workbook_row in sheet.iter_rows()
        for cell in workbook_row
    )


def test_interpretation_controller_synchronizes_interval_selection() -> None:
    controller = make_controller()
    primary = controller.add_interpretation("Primary")
    interval = controller.add_interval(100.0, 140.0, "Reservoir", "Sand A")
    secondary = controller.add_interpretation("Secondary")

    selected = controller.select_interval(primary.interpretation_id, interval.interval_id)

    assert selected is interval
    assert controller.selected_interpretation_id == primary.interpretation_id
    assert controller.selected_interval() is interval

    controller.select_interpretation(secondary.interpretation_id)

    assert controller.selected_interval_id is None
    assert controller.selected_interval() is None


def test_interpretation_controller_normalizes_selection_after_well_switch() -> None:
    controller = make_controller()
    interpretation = controller.add_interpretation("Primary")
    interval = controller.add_interval(100.0, 140.0, "Reservoir", "Sand A")
    controller.select_interval(interpretation.interpretation_id, interval.interval_id)
    session = controller.session
    assert session.project is not None
    second_well = Well("well-two", "Well B")
    session.project.wells[second_well.well_id] = second_well
    session.current_well_id = second_well.well_id
    session.current_dataset_id = None

    controller.normalize_selection()

    assert controller.selected_interpretation_id is None
    assert controller.selected_interval_id is None
