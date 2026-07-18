from openpyxl import Workbook
from PySide6.QtWidgets import QDialogButtonBox

from geoworkbench.services.localization import AppLanguage
from geoworkbench.ui.excel_import_dialog import ExcelImportDialog


def test_excel_dialog_selects_sheet_header_and_index(qapp, tmp_path) -> None:
    source = tmp_path / "logging.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Logging"
    sheet.append(["title", None])
    sheet.append(["DEPTH", "C1"])
    sheet.append([100.0, 1.0])
    workbook.save(source)

    dialog = ExcelImportDialog(source)
    dialog.header_row.setValue(2)
    dialog.index_column.setCurrentText("DEPTH")

    plan = dialog.import_plan()

    assert plan.sheet_name == "Logging"
    assert plan.header_row == 2
    assert plan.index_column == "DEPTH"
    dialog.close()


def test_excel_dialog_builds_composite_time_plan(qapp, tmp_path) -> None:
    source = tmp_path / "time.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Time"
    sheet.append(["DATE", "TIME", "C1"])
    sheet.append(["15.07.2026", "10:00:00", 1.0])
    workbook.save(source)
    dialog = ExcelImportDialog(source)
    dialog.index_column.setCurrentText("DATE")
    dialog.composite_time.setChecked(True)
    dialog.time_column.setCurrentText("TIME")
    dialog.date_format.setText("%d.%m.%Y")
    dialog.timezone.setText("Asia/Oral")

    plan = dialog.import_plan()

    assert plan.index_column == "DATE"
    assert plan.time_column == "TIME"
    assert plan.date_format == "%d.%m.%Y"
    assert plan.timezone == "Asia/Oral"
    dialog.close()


def test_excel_dialog_uses_kazakh_catalog(qapp, tmp_path) -> None:
    source = tmp_path / "logging.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["DEPTH", "C1"])
    sheet.append([100.0, 1.0])
    workbook.save(source)
    dialog = ExcelImportDialog(source, language=AppLanguage.KK)
    buttons = dialog.findChild(QDialogButtonBox)

    assert dialog.windowTitle() == "Excel импорттау — logging.xlsx"
    assert dialog.composite_time.text().endswith("біріктіру")
    assert dialog.timezone.placeholderText().startswith("мысалы")
    assert buttons.button(QDialogButtonBox.StandardButton.Cancel).text() == "Бас тарту"
    dialog.close()
