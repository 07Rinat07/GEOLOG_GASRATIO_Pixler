from __future__ import annotations

from types import SimpleNamespace
import zipfile

from openpyxl import Workbook

from geoworkbench.data.hydrocarbon_interpretation_export import (
    _opus_gasomer_docx,
    _write_docx,
)
from geoworkbench.data.hydrocarbon_interpretation_export_readable import (
    _write_methods_sheet,
    _write_opus_gasomer_sheet,
)


def _minimal_report(**overrides):
    values = {
        "report_profile": "standard",
        "project_name": "Project",
        "well_name": "Well",
        "dataset_name": "Dataset",
        "generated_at": "2026-09-04T00:00:00Z",
        "primary_mnemonic": "TG",
        "threshold": 3.0,
        "methods": (),
        "opus_gasomer": None,
        "candidates": (),
        "depth_unit": "m",
        "manual_intervals": (),
        "warnings": ("internal warning must stay out of client export",),
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def _minimal_opus_section():
    return SimpleNamespace(
        input_curves=(("TOTAL_GAS", "TG"),),
        input_units=(("TOTAL_GAS", "%"),),
        total_gas_lod=0.01,
        working_unit="%",
        profile_id="opus-test",
        profile_version="1",
        profile_status="verified",
        calculation_mode="test",
        interval_source="test",
        formulas=(("Wh", "C2+C3+C4+C5"),),
        intervals=(),
        provenance=("published formula source",),
        source_workbook_sha256="abc123",
        errata=("source workbook correction",),
        warnings=("internal opus warning must stay out of client export",),
    )


def _workbook_values(workbook: Workbook) -> tuple[str, ...]:
    values: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
    return tuple(values)


def test_docx_omits_methodology_limitations_but_keeps_report_content(tmp_path) -> None:
    target = tmp_path / "report.docx"
    _write_docx(target, _minimal_report(), None)

    with zipfile.ZipFile(target) as package:
        document_xml = package.read("word/document.xml").decode("utf-8")

    assert "Ограничения методики" not in document_xml
    assert "internal warning must stay out of client export" not in document_xml
    assert "Методы и доступность" in document_xml
    assert "Интервалы, подтверждённые геологом" in document_xml


def test_methods_sheet_omits_methodology_limitations_and_internal_warnings() -> None:
    method = SimpleNamespace(
        method="Pixler",
        available=True,
        available_mnemonics=("C1", "C2"),
        calculation="ratio",
        source="published source",
    )
    workbook = Workbook()
    _write_methods_sheet(workbook, _minimal_report(methods=(method,)))
    values = _workbook_values(workbook)

    assert "Ограничения методики" not in values
    assert "internal warning must stay out of client export" not in values
    assert "Pixler" in values
    assert "published source" in values


def test_opus_docx_keeps_provenance_and_errata_without_client_warnings() -> None:
    section = _minimal_opus_section()
    report = _minimal_report(report_profile="opus", opus_gasomer=section)
    text = "\n".join(_opus_gasomer_docx(report))

    assert "Происхождение формул" in text
    assert "Происхождение формул и ограничения" not in text
    assert "published formula source" in text
    assert "source workbook correction" in text
    assert "internal opus warning must stay out of client export" not in text


def test_opus_xlsx_keeps_provenance_and_errata_without_client_warnings() -> None:
    section = _minimal_opus_section()
    report = _minimal_report(report_profile="opus", opus_gasomer=section)
    workbook = Workbook()
    _write_opus_gasomer_sheet(workbook, report)
    values = _workbook_values(workbook)

    assert "Происхождение формул" in values
    assert "QC и ограничения" not in values
    assert "published formula source" in values
    assert "source workbook correction" in values
    assert "internal opus warning must stay out of client export" not in values
