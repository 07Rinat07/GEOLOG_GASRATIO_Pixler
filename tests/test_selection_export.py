import csv
from datetime import datetime
import zipfile

import numpy as np

from geoworkbench.data.selection_export import (
    EXCEL_DECIMAL_NUMBER_FORMAT,
    export_selection_excel,
    export_selection_text,
)
from openpyxl import load_workbook

from geoworkbench.domain.models import (
    CurveData,
    CurveMetadata,
    Dataset,
    DatasetIndex,
    DatasetKind,
    DepthDomain,
    IndexRole,
    IndexType,
)


def make_dataset() -> Dataset:
    dataset = Dataset(
        "dataset",
        "Well data",
        DatasetKind.GTI,
        DepthDomain.MD,
        np.array([100.0, 101.0, 102.0, 103.0]),
    )
    dataset.curves["c1"] = CurveData(
        CurveMetadata("c1", "C1", "C1", "%", "Methane", dataset.dataset_id),
        np.array([1.0, 2.0, 3.0, 4.0]),
    )
    dataset.curves["rop"] = CurveData(
        CurveMetadata("rop", "ROP", "ROP", "m/h", "ROP", dataset.dataset_id),
        np.array([10.0, 20.0, 30.0, 40.0]),
    )
    return dataset


def test_text_export_contains_only_selected_interval_and_curves(tmp_path) -> None:
    target = tmp_path / "selection.txt"

    export_selection_text(make_dataset(), target, ["c1", "rop"], 101.0, 102.0, delimiter="\t")

    with target.open(encoding="utf-8", newline="") as stream:
        rows = list(csv.reader(stream, delimiter="\t"))
    assert rows == [
        ["DEPTH [m]", "C1 [%]", "ROP [m/h]"],
        ["101", "2", "20"],
        ["102", "3", "30"],
    ]


def test_text_export_writes_small_values_without_scientific_notation(tmp_path) -> None:
    dataset = make_dataset()
    dataset.curves["c1"].values[1] = 5.2e-5
    target = tmp_path / "small-values.csv"

    export_selection_text(dataset, target, ["c1"], 101.0, 101.0, delimiter=",")

    with target.open(encoding="utf-8", newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows[1] == ["101", "0.000052"]


def test_excel_export_is_valid_openxml_with_data_and_metadata_sheets(tmp_path) -> None:
    target = tmp_path / "selection.xlsx"

    export_selection_excel(make_dataset(), target, ["c1"], 101.0, 102.0)

    with zipfile.ZipFile(target) as archive:
        assert archive.testzip() is None
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
    assert 'name="Data"' in workbook_xml
    assert 'name="Parameters"' in workbook_xml
    assert 'name="Metadata"' in workbook_xml

    workbook = load_workbook(target, data_only=True)
    assert workbook["Data"]["A1"].value == "Глубина по стволу\nDEPTH\n[m]"
    assert workbook["Data"]["B1"].value == "Метан\nC1\n[%]"
    assert workbook["Parameters"]["B1"].value == "Понятное название"
    assert dict(workbook["Metadata"].values)["Набор данных"] == "Well data"


def test_excel_export_keeps_small_values_numeric_with_decimal_format(tmp_path) -> None:
    dataset = make_dataset()
    dataset.curves["c1"].values[1] = 5.2e-5
    target = tmp_path / "small-values.xlsx"

    export_selection_excel(dataset, target, ["c1"], 101.0, 101.0)

    cell = load_workbook(target, data_only=True)["Data"]["B2"]
    assert cell.value == 0.000052
    assert cell.data_type == "n"
    assert cell.number_format == EXCEL_DECIMAL_NUMBER_FORMAT


def test_excel_export_formats_datetime_index_as_excel_date_and_time(tmp_path) -> None:
    dataset = make_dataset()
    dataset.add_index(
        DatasetIndex(
            "datetime",
            "DATETIME",
            IndexType.DATETIME,
            IndexRole.TIME,
            None,
            np.array(
                [
                    "2026-07-18T08:15:30.125",
                    "2026-07-18T08:15:31.250",
                    "2026-07-18T08:15:32.375",
                    "2026-07-18T08:15:33.500",
                ],
                dtype="datetime64[ns]",
            ),
            timezone="UTC",
        ),
        make_active=True,
    )
    target = tmp_path / "time-selection.xlsx"

    export_selection_excel(dataset, target, ["c1"], 101.0, 102.0)

    workbook = load_workbook(target, data_only=True)
    data_sheet = workbook["Data"]
    metadata_sheet = workbook["Metadata"]
    assert data_sheet["A1"].value == "Дата и время\nDATETIME\n[UTC]"
    assert data_sheet["A2"].value == datetime(2026, 7, 18, 8, 15, 31, 250000)
    assert data_sheet["A3"].value == datetime(2026, 7, 18, 8, 15, 32, 375000)
    assert data_sheet["A2"].number_format == "yyyy-mm-dd hh:mm:ss.000"
    assert dict(metadata_sheet.values)["Часовой пояс источника DATETIME"] == "UTC"


def test_excel_depth_export_includes_formatted_secondary_datetime_index(tmp_path) -> None:
    dataset = make_dataset()
    dataset.add_index(
        DatasetIndex(
            "datetime",
            "DATETIME",
            IndexType.DATETIME,
            IndexRole.TIME,
            None,
            np.array(
                [
                    "2026-07-18T08:00:00",
                    "2026-07-18T08:00:01",
                    "2026-07-18T08:00:02",
                    "2026-07-18T08:00:03",
                ],
                dtype="datetime64[ns]",
            ),
            timezone="Asia/Oral",
        )
    )
    target = tmp_path / "depth-with-time.xlsx"

    export_selection_excel(dataset, target, ["c1"], 101.0, 102.0)

    sheet = load_workbook(target, data_only=True)["Data"]
    assert [cell.value for cell in sheet[1]] == [
        "Глубина по стволу\nDEPTH\n[m]",
        "Дата и время\nDATETIME\n[UTC]",
        "Метан\nC1\n[%]",
    ]
    assert sheet["A2"].value == 101.0
    assert sheet["B2"].value == datetime(2026, 7, 18, 8, 0, 1)
    assert sheet["B2"].number_format == "yyyy-mm-dd hh:mm:ss.000"


def test_selection_export_distinguishes_zero_missing_and_unavailable(tmp_path) -> None:
    dataset = make_dataset()
    dataset.curves["c1"].values[:] = np.array([0.0, np.nan, 3.0, 4.0])

    csv_target = tmp_path / "coverage.csv"
    xlsx_target = tmp_path / "coverage.xlsx"
    export_selection_text(
        dataset,
        csv_target,
        ["c1"],
        100.0,
        102.0,
        delimiter=",",
        unavailable_mnemonics=("H2S",),
    )
    export_selection_excel(
        dataset,
        xlsx_target,
        ["c1"],
        100.0,
        102.0,
        unavailable_mnemonics=("H2S",),
    )

    rows = list(csv.reader(csv_target.open(encoding="utf-8")))
    assert rows[1][1:] == ["0", "#N/A"]
    assert rows[2][1:] == ["", "#N/A"]

    workbook = load_workbook(xlsx_target, data_only=True)
    data = workbook["Data"]
    assert data[2][1].value == 0.0
    assert data[3][1].value is None
    assert data[2][2].value == "#N/A"
    parameters = workbook["Parameters"]
    assert parameters[3][10].value == "available"
    assert parameters[3][11].value == 2
    assert parameters[3][12].value == 1
    assert parameters[3][13].value == 1
    assert parameters[4][10].value == "unavailable"
    assert parameters[4][14].value == 0.0
