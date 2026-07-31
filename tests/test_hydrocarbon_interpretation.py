from __future__ import annotations

import zipfile

import numpy as np
from openpyxl import load_workbook

from geoworkbench.data.hydrocarbon_interpretation_export import (
    HydrocarbonInterpretationExportError,
    export_hydrocarbon_interpretation_docx,
    export_hydrocarbon_interpretation_xlsx,
)
from geoworkbench.domain.models import (
    CurveData,
    CurveMetadata,
    CuttingsSample,
    Dataset,
    DatasetKind,
    DepthDomain,
)
from geoworkbench.project.interpretation_controller import InterpretationController
from geoworkbench.project.session import ProjectSession
from geoworkbench.services.hydrocarbon_interpretation import (
    build_hydrocarbon_interpretation_report,
    hydrocarbon_interpretation_html,
)
from geoworkbench.services.localization import AppLanguage


def _session() -> ProjectSession:
    depth = np.arange(1_000.0, 1_100.0)
    dataset = Dataset(
        "well-log",
        "Whole <well> log",
        DatasetKind.GTI,
        DepthDomain.MD,
        depth,
    )
    normalized_gas = np.ones(depth.shape)
    normalized_gas[40:43] = (80.0, 120.0, 90.0)
    for mnemonic, values, unit in (
        ("C1_NORM", normalized_gas, "normalized gas units"),
        ("WH", np.full(depth.shape, 12.0), "%"),
        ("C1_C2", np.full(depth.shape, 8.0), "ratio"),
        ("DEXP", np.full(depth.shape, 1.4), "dimensionless"),
    ):
        dataset.curves[mnemonic] = CurveData(
            CurveMetadata(
                mnemonic,
                mnemonic,
                mnemonic,
                unit,
                mnemonic,
                dataset.dataset_id,
                "calculation:test",
            ),
            values,
        )
    variation = np.resize(np.array([-0.30, -0.15, 0.0, 0.15, 0.30]), depth.shape)
    gas_components = {
        "C1": np.full(depth.shape, 90.0),
        "C2": 5.0 + variation,
        "C3": 3.0 + variation / 2.0,
        "IC4": np.full(depth.shape, 1.0),
        "NC4": np.full(depth.shape, 1.0),
        "IC5": np.full(depth.shape, 0.5),
        "NC5": np.full(depth.shape, 0.5),
    }
    for values, high_value in (
        (gas_components["C2"], 30.0),
        (gas_components["C3"], 20.0),
        (gas_components["IC4"], 5.0),
        (gas_components["NC4"], 5.0),
        (gas_components["IC5"], 2.5),
        (gas_components["NC5"], 2.5),
    ):
        values[40:43] = high_value
    for mnemonic, values in gas_components.items():
        dataset.curves[mnemonic] = CurveData(
            CurveMetadata(
                mnemonic,
                mnemonic,
                mnemonic,
                "%",
                mnemonic,
                dataset.dataset_id,
                "source:test",
            ),
            values,
        )
    session = ProjectSession()
    session.project.name = "=Project formula"
    session.add_dataset(dataset, "Well <A>")
    interpretation = InterpretationController(session)
    interpretation.add_interpretation("Geologist")
    interpretation.add_interval(
        1_039.5,
        1_043.5,
        "hydrocarbon show",
        "+Confirmed",
        comment="Check DST",
    )
    return session


def test_report_detects_relative_anomaly_and_keeps_manual_intervals_separate() -> None:
    report = build_hydrocarbon_interpretation_report(_session(), threshold=3.0)

    assert report.primary_mnemonic == "C1_NORM"
    assert len(report.candidates) == 1
    candidate = report.candidates[0]
    assert candidate.top_depth <= 1_040.0
    assert candidate.bottom_depth >= 1_042.0
    assert candidate.sample_count == 3
    assert candidate.anomaly_strength == "high"
    assert candidate.fluid_hypothesis == "heavy_or_residual_oil"
    assert candidate.wetness_robust_z is not None
    assert candidate.wetness_robust_z > 2.0
    assert ("WH", 12.0) in candidate.metrics
    assert any("context means: WH=12" in evidence for evidence in candidate.evidence)
    assert len(report.manual_intervals) == 1
    assert report.manual_intervals[0].label == "+Confirmed"
    assert any("не заключение" in warning for warning in report.warnings)

    html = hydrocarbon_interpretation_html(report, AppLanguage.RU)
    assert "background: #ffffff" in html
    assert "td { background: #ffffff; }" in html
    assert "Well &lt;A&gt;" in html
    assert "Перспективные интервалы" in html
    assert "Кандидатные интервалы" not in html
    assert "page-break-before: always" in html
    assert "тяжёлая или остаточная нефть" in html
    assert "Check DST" in html


def test_report_exports_openable_xlsx_and_docx(tmp_path) -> None:
    session = _session()
    dataset = session.current_dataset
    assert dataset is not None
    report = build_hydrocarbon_interpretation_report(session)
    xlsx_path = export_hydrocarbon_interpretation_xlsx(
        report,
        dataset,
        tmp_path / "interpretation.xlsx",
    )
    docx_path = export_hydrocarbon_interpretation_docx(
        report,
        tmp_path / "interpretation.docx",
        dataset=dataset,
    )

    workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        assert workbook.sheetnames == [
            "Интерпретация УВ",
            "Методика",
            "Данные по глубине",
        ]
        main = workbook["Интерпретация УВ"]
        assert main["B2"].value == "'=Project formula"
        assert main["A5"].value == "Перспективных УВ-интервалов"
        headers = [main.cell(9, column).value for column in range(1, 24)]
        assert "Абсолютный газ по компонентам: мин / среднее / макс" in headers
        assert "Точек выше порога" not in headers
        assert not any("Медиана" in str(value) for value in headers)
        assert not any("Фон" in str(value) for value in headers)
        assert main["F10"].value == "Подтвержден геологом"
        assert "Кандидат" not in str(main["F10"].value)
        assert "C1" in str(main["R10"].value)
        assert "IC4" in str(main["R10"].value)
        assert "NC5" in str(main["R10"].value)
        assert workbook["Данные по глубине"].max_row == dataset.depth.size + 1
    finally:
        workbook.close()
    with zipfile.ZipFile(docx_path) as package:
        assert package.testzip() is None
        document = package.read("word/document.xml").decode("utf-8")
        assert "Перспективные интервалы" in document
        assert "Кандидатные интервалы" not in document
        assert "тяжёлая или остаточная нефть" in document
        assert "Абсолютный газ: мин / среднее / макс" in document
        assert "Точек выше порога" not in document
        assert "Медиана" not in document
        assert "Фон" not in document
        assert "IC4" in document and "NC5" in document
        assert "Check DST" in document


def test_report_falls_back_from_sparse_normalized_gas_to_total_gas() -> None:
    session = _session()
    dataset = session.current_dataset
    assert dataset is not None
    dataset.curves["C1_NORM"].values[:] = np.nan
    total_gas = np.ones(dataset.depth.shape)
    total_gas[60:63] = (70.0, 110.0, 80.0)
    dataset.curves["TG"] = CurveData(
        CurveMetadata(
            "TG",
            "TG",
            "TG",
            "%",
            "TG",
            dataset.dataset_id,
            "source:test",
        ),
        total_gas,
    )

    report = build_hydrocarbon_interpretation_report(session)

    assert report.primary_mnemonic == "TG"
    assert len(report.candidates) == 1


def test_report_uses_server_normalized_gas_alias_and_discloses_its_origin() -> None:
    session = _session()
    dataset = session.current_dataset
    assert dataset is not None
    server_values = np.ones(dataset.depth.shape)
    server_values[65:68] = (75.0, 115.0, 85.0)
    dataset.curves["server-tg-norm"] = CurveData(
        CurveMetadata(
            "server-tg-norm",
            "NORMALIZED_TOTAL_GAS",
            "NORMALIZED_TOTAL_GAS",
            "normalized gas units",
            "Operator normalized total gas",
            dataset.dataset_id,
            "source:server",
        ),
        server_values,
    )

    report = build_hydrocarbon_interpretation_report(session)

    assert report.primary_mnemonic == "NORMALIZED_TOTAL_GAS"
    assert any(
        "NORMALIZED_TOTAL_GAS" in warning and "из файла/сервера" in warning
        for warning in report.warnings
    )
    assert len(report.candidates) == 1


def test_report_can_interpret_probable_gas_without_claiming_final_fluid_type() -> None:
    session = _session()
    dataset = session.current_dataset
    assert dataset is not None
    dataset.curves["C1"].values[40:43] = 300.0
    for mnemonic in ("C2", "C3", "IC4", "NC4", "IC5", "NC5"):
        dataset.curves[mnemonic].values[40:43] = 0.1

    report = build_hydrocarbon_interpretation_report(session)

    assert report.candidates[0].fluid_hypothesis == "very_light_dry_gas"
    html = hydrocarbon_interpretation_html(report, AppLanguage.RU)
    assert "очень лёгкий сухой газ" in html
    assert "Категория «вода» по mud-gas не назначается" in html


def test_sparse_heavy_components_use_integrated_interval_composition() -> None:
    session = _session()
    dataset = session.current_dataset
    assert dataset is not None
    for mnemonic in ("C2", "C3", "IC4", "NC4", "IC5", "NC5"):
        values = dataset.curve_by_mnemonic(mnemonic).values
        values[40:43] = (0.0, 0.0, 0.001)

    report = build_hydrocarbon_interpretation_report(session)

    candidate = report.candidates[0]
    assert candidate.interval_wetness is not None
    assert candidate.interval_wetness > 0.0
    html = hydrocarbon_interpretation_html(report, AppLanguage.RU)
    assert f"{candidate.interval_wetness:.5f}%" in html
    assert "Точек выше порога" not in html
    assert "Абсолютный газ: мин / среднее / макс" in html
    assert "фон 0.00000" not in html


def test_report_correlates_gas_interpretation_with_overlapping_lba() -> None:
    session = _session()
    well = session.current_well
    assert well is not None
    well.cuttings.append(
        CuttingsSample(
            "lba-overlap",
            1_039.5,
            1_043.5,
            lba_group=4,
            lba_type_id="СБ",
            lba_intensity=4,
            lba_color="ОК — оранжево-коричневый",
        )
    )

    report = build_hydrocarbon_interpretation_report(session)

    candidate = report.candidates[0]
    assert candidate.gas_lba_correlation == "concordant"
    assert candidate.lba_assessments[0].standard.code == "СБ"
    html = hydrocarbon_interpretation_html(report, AppLanguage.RU)
    assert "ЛБА: группа 4" in html
    assert "признаки согласуются" in html


def test_xlsx_export_rejects_mismatched_curve_lengths(tmp_path) -> None:
    session = _session()
    dataset = session.current_dataset
    assert dataset is not None
    report = build_hydrocarbon_interpretation_report(session)
    dataset.curves["WH"].values = dataset.curves["WH"].values[:-1]

    with np.testing.assert_raises_regex(
        HydrocarbonInterpretationExportError,
        "WH",
    ):
        export_hydrocarbon_interpretation_xlsx(
            report,
            dataset,
            tmp_path / "invalid.xlsx",
        )
