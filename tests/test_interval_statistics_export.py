import csv

from openpyxl import load_workbook

from geoworkbench.calculations.interval_statistics import CurveIntervalStatistics
from geoworkbench.data.interval_statistics_export import (
    export_interval_statistics_csv,
    export_interval_statistics_xlsx,
    statistics_tsv,
)
from geoworkbench.services.localization import AppLanguage


def _statistics() -> tuple[CurveIntervalStatistics, ...]:
    return (CurveIntervalStatistics("ROP", "m/h", 3, 1.0, 5.0, 3.0, 4),)


def test_interval_statistics_tsv_is_excel_pasteable() -> None:
    text = statistics_tsv(
        _statistics(), interval_label="Depth: 100–103 m", dataset_name="Well A"
    )

    assert "Dataset\tWell A" in text
    assert "Parameter\tMnemonic\tUnit\tAvailability\tObserved\tZeros\tMissing\tCoverage, %" in text
    assert "ROP\tROP\tm/h\tAvailable\t3\t0\t1\t75" in text


def test_interval_statistics_exports_csv_and_xlsx(tmp_path) -> None:
    csv_path = export_interval_statistics_csv(
        tmp_path / "statistics.csv",
        _statistics(),
        interval_label="Depth: 100–103 m",
        dataset_name="Well A",
    )
    xlsx_path = export_interval_statistics_xlsx(
        tmp_path / "statistics.xlsx",
        _statistics(),
        interval_label="Depth: 100–103 m",
        dataset_name="Well A",
    )

    assert "ROP,ROP,m/h,Available,3,0,1,75.0,1.0,5.0,3.0" in csv_path.read_text(
        encoding="utf-8-sig"
    )
    sheet = load_workbook(xlsx_path, data_only=True).active
    assert sheet["A1"].value == "Dataset"
    assert sheet["B1"].value == "Well A"
    assert sheet["A5"].value == "ROP"
    assert sheet["B5"].value == "ROP"
    assert sheet["H5"].value == 75.0


def test_statistics_csv_and_xlsx_protect_formula_like_text_only(tmp_path) -> None:
    mnemonic = "\n+MNEMONIC"
    statistics = (
        CurveIntervalStatistics(mnemonic, "=UNIT", 3, -5.0, 5.0, 0.0, 4),
    )
    display_names = {mnemonic: "\ufeff@DISPLAY"}
    csv_path = export_interval_statistics_csv(
        tmp_path / "formula-guard.csv",
        statistics,
        interval_label="\n-INTERVAL",
        dataset_name=" \t=DATA()",
        display_names=display_names,
    )
    xlsx_path = export_interval_statistics_xlsx(
        tmp_path / "formula-guard.xlsx",
        statistics,
        interval_label="\n-INTERVAL",
        dataset_name=" \t=DATA()",
        display_names=display_names,
    )

    with csv_path.open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.reader(stream))
    assert rows[0][1] == "' \t=DATA()"
    assert rows[1][1] == "'\n-INTERVAL"
    assert rows[4][0] == "'\ufeff@DISPLAY"
    assert rows[4][1] == "'\n+MNEMONIC"
    assert rows[4][2] == "'=UNIT"
    assert rows[4][8] == "-5.0"

    sheet = load_workbook(xlsx_path, data_only=False).active
    assert sheet["B1"].value == "' \t=DATA()"
    assert sheet["B2"].value == "'\n-INTERVAL"
    assert sheet["A5"].value == "'\ufeff@DISPLAY"
    assert sheet["B5"].value == "'\n+MNEMONIC"
    assert sheet["C5"].value == "'=UNIT"
    assert sheet["I5"].value == -5.0
    assert sheet["I5"].data_type == "n"
    assert all(
        cell.data_type != "f"
        for row in sheet.iter_rows()
        for cell in row
    )


def test_interval_statistics_exports_missing_values_as_empty_cells(tmp_path) -> None:
    missing = (CurveIntervalStatistics("EMPTY", "ppm", 0, float("nan"), float("nan"), float("nan"), 4),)

    text = statistics_tsv(
        missing, interval_label="Depth: 100–103 m", dataset_name="Well A"
    )
    xlsx_path = export_interval_statistics_xlsx(
        tmp_path / "missing.xlsx",
        missing,
        interval_label="Depth: 100–103 m",
        dataset_name="Well A",
    )

    assert "EMPTY\tEMPTY\tppm\tAvailable\t0\t0\t4\t0\t\t\t" in text
    sheet = load_workbook(xlsx_path, data_only=True).active
    assert sheet["I5"].value is None
    assert sheet["J5"].value is None
    assert sheet["K5"].value is None


def test_interval_statistics_xlsx_contains_localized_name_and_mnemonic(tmp_path) -> None:
    xlsx_path = export_interval_statistics_xlsx(
        tmp_path / "localized.xlsx",
        _statistics(),
        interval_label="Глубина: 100–103 м",
        dataset_name="Скважина A",
        display_names={"ROP": "Скорость бурения"},
        language=AppLanguage.RU,
    )

    sheet = load_workbook(xlsx_path, data_only=True).active
    assert sheet["A1"].value == "Набор данных"
    assert sheet["A4"].value == "Параметр"
    assert sheet["B4"].value == "Мнемоника"
    assert sheet["A5"].value == "Скорость бурения"
    assert sheet["B5"].value == "ROP"
