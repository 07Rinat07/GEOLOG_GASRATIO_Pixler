from __future__ import annotations

import numpy as np
from openpyxl import load_workbook  # type: ignore[import-untyped]

from geoworkbench.data.hydrocarbon_interpretation_export_readable import (
    export_readable_hydrocarbon_interpretation_xlsx,
)
from geoworkbench.domain.models import (
    CurveData,
    CurveMetadata,
    Dataset,
    DatasetKind,
    DepthDomain,
)
from geoworkbench.project.interpretation_calculation_controller import (
    NormalizedGasCalculationMode,
)
from geoworkbench.project.session import ProjectSession
from geoworkbench.services.hydrocarbon_interpretation import (
    build_hydrocarbon_interpretation_report,
    hydrocarbon_interpretation_html,
)
from geoworkbench.services.localization import AppLanguage


def _add_curve(
    dataset: Dataset,
    mnemonic: str,
    values: np.ndarray,
    unit: str,
    provenance: str = "source:test",
) -> None:
    dataset.curves[mnemonic] = CurveData(
        CurveMetadata(
            mnemonic,
            mnemonic,
            mnemonic,
            unit,
            mnemonic,
            dataset.dataset_id,
            provenance,
        ),
        np.asarray(values, dtype=np.float64),
    )


def _session_with_zero_heavy_components() -> ProjectSession:
    depth = np.arange(1_000.0, 1_040.0)
    dataset = Dataset(
        "readable-interpretation",
        "Readable interpretation",
        DatasetKind.GTI,
        DepthDomain.MD,
        depth,
    )
    normalized = np.ones(depth.shape)
    normalized[20:23] = (10.0, 12.0, 11.0)
    raw_total = np.full(depth.shape, 0.2)
    raw_total[20:23] = (3.0, 4.0, 5.0)
    _add_curve(
        dataset,
        "TG_NORM_CALC",
        normalized,
        "normalized gas units",
        "calculation:test",
    )
    _add_curve(dataset, "TG_CALC", raw_total, "%abs")
    _add_curve(dataset, "C1", np.full(depth.shape, 10.0), "%")
    for mnemonic in ("C2", "C3", "C4", "C5"):
        _add_curve(dataset, mnemonic, np.zeros(depth.shape), "%")
    _add_curve(
        dataset,
        "DEXP",
        np.linspace(0.7, 1.1, depth.size),
        "dimensionless",
    )

    session = ProjectSession()
    session.add_dataset(dataset, "Well readable")
    session.dirty = False
    return session


def test_html_replaces_ambiguous_zero_wetness_with_explicit_gas_readings() -> None:
    session = _session_with_zero_heavy_components()
    report = build_hydrocarbon_interpretation_report(
        session,
        threshold=3.0,
        normalized_gas_mode=NormalizedGasCalculationMode.LOCAL,
    )

    html = hydrocarbon_interpretation_html(report, AppLanguage.RU)

    assert report.candidates
    assert "Показания газа по интервалам" in html
    assert "Исходный общий газ" in html
    assert "C2-C5 в интервале и на фоне не зарегистрированы выше нуля" in html
    assert "Относительная доля C2–C5: интервал 0.00000%" not in html


def test_readable_xlsx_keeps_interpretation_and_gas_statistics_on_main_sheet(
    tmp_path,
) -> None:
    session = _session_with_zero_heavy_components()
    dataset = session.current_dataset
    assert dataset is not None
    report = build_hydrocarbon_interpretation_report(
        session,
        threshold=3.0,
        normalized_gas_mode=NormalizedGasCalculationMode.LOCAL,
    )
    target = tmp_path / "interpretation.xlsx"

    export_readable_hydrocarbon_interpretation_xlsx(report, dataset, target)

    workbook = load_workbook(target, data_only=False)
    try:
        assert workbook.sheetnames[:2] == ["Интерпретация УВ", "Методика"]
        assert "Candidate intervals" not in workbook.sheetnames
        assert workbook["Данные по глубине"].sheet_state == "hidden"
        sheet = workbook["Интерпретация УВ"]
        headers = [sheet.cell(9, column).value for column in range(1, 28)]
        assert "Предварительная интерпретация" in headers
        assert "Фон исходного газа" in headers
        assert "Мин исходного газа" in headers
        assert "Макс исходного газа" in headers
        assert sheet["F10"].value == "Кандидат УВ-пласта"
        assert isinstance(sheet["G10"].value, str) and sheet["G10"].value
        assert sheet["I10"].value == "TG_CALC [%abs]"
        assert sheet["J10"].value == 0.2
        assert sheet["K10"].value == 3.0
        assert sheet["L10"].value == 4.0
        assert sheet["M10"].value == 4.0
        assert sheet["N10"].value == 5.0
        assert "выше нуля не зарегистрированы" in str(sheet["V10"].value)
        assert "0 — реальное нулевое измерение" in str(sheet["A7"].value)
    finally:
        workbook.close()
