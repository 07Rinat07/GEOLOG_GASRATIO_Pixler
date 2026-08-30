from __future__ import annotations

import zipfile
from html import unescape

import fitz
import numpy as np
from openpyxl import load_workbook

from geoworkbench.data.hydrocarbon_interpretation_export import (
    export_hydrocarbon_interpretation_docx,
    export_hydrocarbon_interpretation_xlsx,
)
from geoworkbench.domain.models import Dataset, DatasetKind, DepthDomain
from geoworkbench.project.session import ProjectSession
from geoworkbench.project.interpretation_calculation_controller import (
    InterpretationCalculationController,
)
from geoworkbench.printing.hydrocarbon_interpretation_report import (
    export_hydrocarbon_interpretation_pdf,
)
from geoworkbench.printing.hydrocarbon_interpretation_chart_front import (
    hydrocarbon_interpretation_html_with_front_chart,
)
from geoworkbench.services.hydrocarbon_interpretation import (
    build_opus_interpretation_report,
    fluid_hypothesis_label,
    hydrocarbon_interpretation_html,
)
from geoworkbench.services.localization import AppLanguage


def _gasomer_session() -> ProjectSession:
    depth = np.arange(0.0, 100.2, 0.2)
    total = np.full(depth.shape, 0.01)
    event = (depth >= 40.0) & (depth <= 42.0)
    total[event] = 0.20
    shares = {"C1": 3.0, "C2": 2.0, "C3": 1.0, "C4": 1.0, "C5": 1.0}
    dataset = Dataset(
        "opus-gasomer-report",
        "OPUS Gasomer report source",
        DatasetKind.GTI,
        DepthDomain.MD,
        depth,
    )
    for mnemonic, share in shares.items():
        dataset.upsert_curve(
            mnemonic,
            total * share / 9.0,
            unit="%vol",
            description=f"Source {mnemonic}",
            provenance="source:test",
        )
    dataset.upsert_curve(
        "TOTAL_GAS",
        total,
        unit="%vol",
        description="Independent synchronous total gas",
        provenance="source:test",
    )
    session = ProjectSession()
    session.add_dataset(dataset, "Gasomer well")
    return session


def test_gasomer_report_stores_detector_votes_qc_and_provenance(qapp) -> None:
    session = _gasomer_session()
    report = build_opus_interpretation_report(
        session,
        total_gas_lod=0.001,
    )

    section = report.opus_gasomer
    assert section is not None
    assert section.profile_id == "opus-gasomer-total-gas-workbook"
    assert section.interval_source == "opus-gasomer-local-detector"
    assert dict(section.input_curves)["TOTAL_GAS"] == "TOTAL_GAS"
    assert section.total_gas_lod == 0.001
    assert len(section.intervals) == 1
    interval = section.intervals[0]
    assert interval.class_code == 2
    assert interval.class_label == "Нефть"
    assert interval.support_fraction == 1.0
    assert interval.valid_rows == interval.total_rows
    assert interval.background_median == 0.01
    assert interval.max_contrast == 20.0
    assert len(interval.indicators) == 5
    assert {item.class_code for item in interval.indicators} <= {1, 2}
    assert all(item.available_rows == item.total_rows for item in interval.indicators)
    assert dict(section.formulas)["OPUS_GM_5"].endswith("* 100")
    assert len(section.source_workbook_sha256) == 64
    assert any("AB2>AB5=250000" in item for item in section.errata)
    assert any("AB2>=250000" in item for item in section.errata)

    html = hydrocarbon_interpretation_html(report, AppLanguage.RU)
    readable_html = unescape(html)
    assert "ОПУС Газомер — пять показателей и голоса" in readable_html
    assert "GM_5=(p2×p3×p4×p5/p1)" in readable_html
    assert "20.000" in readable_html
    assert "AB2>AB5=250000" in readable_html
    assert "AB2>=250000" in readable_html
    dataset = session.current_dataset
    assert dataset is not None
    print_html = hydrocarbon_interpretation_html_with_front_chart(
        report,
        dataset,
        AppLanguage.RU,
        print_layout=True,
    )
    assert "ОПУС Газомер — пять показателей и голоса" in print_html
    assert "класс 2 — Нефть" in print_html


def test_gasomer_snapshot_is_exported_without_recalculation(tmp_path, qapp) -> None:
    session = _gasomer_session()
    dataset = session.current_dataset
    assert dataset is not None
    report = build_opus_interpretation_report(session, total_gas_lod=0.001)

    xlsx = export_hydrocarbon_interpretation_xlsx(
        report,
        dataset,
        tmp_path / "gasomer.xlsx",
    )
    workbook = load_workbook(xlsx, read_only=True, data_only=False)
    try:
        assert "ОПУС Газомер" in workbook.sheetnames
        values = tuple(
            str(cell.value or "")
            for row in workbook["ОПУС Газомер"].iter_rows()
            for cell in row
        )
        assert any("opus-gasomer-total-gas-workbook" in value for value in values)
        assert any("((p2 * p3 * p4 * p5) / p1)" in value for value in values)
        assert any("Нефть" == value for value in values)
        assert any("available:" in value for value in values)
        assert any("AB2>AB5=250000" in value for value in values)
        assert any("AB2>=250000" in value for value in values)
    finally:
        workbook.close()

    docx = export_hydrocarbon_interpretation_docx(
        report,
        tmp_path / "gasomer.docx",
        dataset=dataset,
    )
    with zipfile.ZipFile(docx) as package:
        document = unescape(package.read("word/document.xml").decode("utf-8"))
    assert "ОПУС Газомер — пять показателей и голоса" in document
    assert "opus-gasomer-total-gas-workbook" in document
    assert "класс 2 — Нефть" in document
    assert "SHA-256 книги" in document
    assert "AB2>AB5=250000" in document
    assert "AB2>=250000" in document

    pdf = export_hydrocarbon_interpretation_pdf(
        report,
        tmp_path / "gasomer.pdf",
        dataset=dataset,
    )
    with fitz.open(pdf) as document_pdf:
        pdf_text = "\n".join(page.get_text() for page in document_pdf)
    assert "ОПУС Газомер" in pdf_text
    assert "opus-gasomer-total-gas-workbook" in pdf_text
    assert "Нефть" in pdf_text
    assert "AB2>=250000" in pdf_text


def test_missing_lod_is_explicit_and_does_not_run_gasomer_detector() -> None:
    report = build_opus_interpretation_report(_gasomer_session())

    section = report.opus_gasomer
    assert section is not None
    assert section.total_gas_lod is None
    assert section.interval_source == "existing-opus-report-candidates"
    assert not section.intervals
    assert any("детектор ОПУС Газомер не запускался" in item for item in section.warnings)


def test_gasomer_class_replaces_ambiguous_historical_headline() -> None:
    session = _gasomer_session()
    InterpretationCalculationController(session).calculate_opus_curves()

    report = build_opus_interpretation_report(session)

    assert len(report.candidates) == 1
    candidate = report.candidates[0]
    assert candidate.fluid_hypothesis == "opus_gasomer_oil"
    assert "ОПУС Газомер: класс 2 — нефть" in fluid_hypothesis_label(
        candidate,
        AppLanguage.RU,
    )
    assert not candidate.fluid_hypothesis.startswith("opus_fallback__")
    assert any(
        "OPUS Gasomer primary result: class=2" in item
        and "valid synchronous rows=" in item
        for item in candidate.evidence
    )


def test_ambiguous_total_gas_uses_explicit_component_sum_policy() -> None:
    session = _gasomer_session()
    dataset = session.current_dataset
    assert dataset is not None
    total = dataset.curve_by_mnemonic("TOTAL_GAS")
    assert total is not None
    dataset.upsert_curve(
        "TG_CALC",
        total.values * 1.01,
        unit="%abs",
        description="Calculated component sum",
        provenance="source:test",
    )
    InterpretationCalculationController(session).calculate_opus_curves()

    report = build_opus_interpretation_report(session)

    section = report.opus_gasomer
    assert section is not None
    assert dict(section.input_curves)["TOTAL_GAS"] == "Σ(C1–C5)"
    assert any(
        "несколько равноправных каналов TotalGas" in warning
        and "синхронной суммой C1–C5" in warning
        for warning in section.warnings
    )
