from __future__ import annotations

import zipfile

import numpy as np
from openpyxl import load_workbook

from geoworkbench.data.hydrocarbon_interpretation_export import (
    export_hydrocarbon_interpretation_docx,
    export_hydrocarbon_interpretation_xlsx,
)
from geoworkbench.calculations.gas_ratio import (
    OPUS_SCREENING_PROFILE_ID,
    calculate_basic_ratios,
    calculate_opus_report_curves,
    calculate_opus_screening,
)
from geoworkbench.domain.models import Dataset, DatasetKind, DepthDomain
from geoworkbench.project.interpretation_calculation_controller import (
    InterpretationCalculationController,
)
from geoworkbench.project.session import ProjectSession
from geoworkbench.printing.hydrocarbon_interpretation_chart import (
    _panel_curves as screen_panel_curves,
)
from geoworkbench.printing.hydrocarbon_interpretation_pdf_chart import (
    _panel_curves as pdf_panel_curves,
)
from geoworkbench.services.opus_interpretation import (
    _classify_opus_interval,
    build_opus_interpretation_report,
)
from geoworkbench.services.hydrocarbon_interpretation import (
    fluid_hypothesis_label,
    hydrocarbon_interpretation_html,
)
from geoworkbench.services.localization import AppLanguage


def test_opus_reproduces_published_c1_c5_worked_example() -> None:
    curves = calculate_opus_screening(
        np.array([98.315]),
        np.array([1.186]),
        np.array([0.206]),
        np.array([0.172]),
        np.array([0.120]),
    )

    assert OPUS_SCREENING_PROFILE_ID == "opus-lukyanov-c1-c5-relative-1987-1997"
    np.testing.assert_allclose(curves["OPUS3"].values, [60.18], rtol=2e-4)
    np.testing.assert_allclose(curves["OPUS4"].values, [6.2875], rtol=2e-3)
    np.testing.assert_allclose(curves["OPUS_K1_3"].values, [8.01], rtol=2e-3)
    # The source prints 0.09 after low-precision intermediate rounding.
    np.testing.assert_allclose(curves["OPUS_1_5"].values, [0.09], atol=0.01)


def test_opus_published_band_intersection_rejects_conflicting_control_vector() -> None:
    class_code, agreement, compatible = _classify_opus_interval(
        {
            "OPUS3": 8.4183673469,
            "OPUS4": 1.5153850661,
            "OPUS_K1_3": 384.5243513883,
            "OPUS_1_5": 11.8906760968,
        }
    )

    assert class_code == 0
    assert agreement == 1.0
    assert dict(compatible)["OPUS_K1_3"] == (5,)


def test_opus_published_band_intersection_accepts_unambiguous_oil() -> None:
    class_code, agreement, _ = _classify_opus_interval(
        {"OPUS3": 1.0, "OPUS4": 0.5, "OPUS_K1_3": 1000.0, "OPUS_1_5": 2000.0}
    )

    assert class_code == 2
    assert agreement == 1.0


def test_opus_published_band_intersection_rejects_out_of_profile_indicator() -> None:
    class_code, agreement, compatible = _classify_opus_interval(
        {"OPUS3": 0.1, "OPUS4": 0.01, "OPUS_K1_3": 0.0, "OPUS_1_5": 0.0}
    )

    assert class_code == 0
    assert agreement == 0.5
    assert dict(compatible)["OPUS_K1_3"] == ()


def test_opus_report_is_separate_and_does_not_double_count_c4_c5() -> None:
    source = {
        "C1": np.array([5.719]),
        "C2": np.array([0.069]),
        "C3": np.array([0.012]),
        "IC4": np.array([0.004]),
        "NC4": np.array([0.006]),
        "IC5": np.array([0.003]),
        "NC5": np.array([0.004]),
        "C4": np.array([99.0]),
        "C5": np.array([99.0]),
    }

    standard = calculate_basic_ratios(source)
    curves = calculate_opus_report_curves(source)

    assert not any(name.startswith("OPUS") for name in standard)
    assert {"OPUS3", "OPUS4", "OPUS_K1_3", "OPUS_1_5"} <= curves.keys()
    assert "OPUS5_REF" not in curves
    np.testing.assert_allclose(curves["OPUS3"].values, [60.1449474], rtol=1e-7)


def test_opus_preserves_invalid_rows_as_nan() -> None:
    curves = calculate_opus_screening(
        np.array([80.0, np.nan, 0.0]),
        np.array([10.0, 10.0, 0.0]),
        np.array([5.0, 5.0, 0.0]),
        np.array([3.0, 3.0, 0.0]),
        np.array([2.0, 2.0, 0.0]),
    )

    for result in curves.values():
        assert np.isfinite(result.values[0])
        assert np.isnan(result.values[1])
        assert np.isnan(result.values[2])


def test_opus_handles_full_5000_m_well_at_point_two_metre_step() -> None:
    depth = np.arange(0.0, 5000.0 + 0.2, 0.2)
    curves = calculate_opus_screening(
        np.full(depth.shape, 80.0),
        np.full(depth.shape, 10.0),
        np.full(depth.shape, 5.0),
        np.full(depth.shape, 3.0),
        np.full(depth.shape, 2.0),
    )

    assert depth.size == 25_001
    assert all(result.values.shape == depth.shape for result in curves.values())
    assert all(np.all(np.isfinite(result.values)) for result in curves.values())


def _opus_session(unit: str, scale: float) -> ProjectSession:
    depth = np.arange(0.0, 100.0, 1.0)
    dataset = Dataset(
        "opus-dataset",
        f"OPUS source {unit}",
        DatasetKind.GTI,
        DepthDomain.MD,
        depth,
    )
    components = {
        "C1": np.full(depth.shape, 0.10 * scale),
        "C2": np.full(depth.shape, 0.02 * scale),
        "C3": np.full(depth.shape, 0.01 * scale),
        "C4": np.full(depth.shape, 0.005 * scale),
        "C5": np.full(depth.shape, 0.005 * scale),
    }
    for values in components.values():
        values[40:43] *= 10.0
    for mnemonic, values in components.items():
        dataset.upsert_curve(
            mnemonic,
            values,
            unit=unit,
            description=f"Source {mnemonic}",
            provenance="source:test",
        )
    session = ProjectSession()
    session.add_dataset(dataset, "OPUS well")
    return session


def test_separate_opus_controller_converts_ppm_to_percent_and_keeps_sources() -> None:
    session = _opus_session("ppm", 10_000.0)
    dataset = session.current_dataset
    assert dataset is not None
    source = dataset.curve_by_mnemonic("C1")
    assert source is not None
    original = source.values.copy()

    result = InterpretationCalculationController(session).calculate_opus_curves()

    assert not result.issues
    assert "OPUS_C1_PCT" in result.created
    np.testing.assert_allclose(dataset.curve_by_mnemonic("OPUS_C1_PCT").values, original * 1e-4)
    np.testing.assert_array_equal(source.values, original)
    assert source.metadata.unit == "ppm"
    assert dataset.curve_by_mnemonic("OPUS_C1_PCT").metadata.unit == "%vol"


def test_ppm_and_percent_inputs_produce_the_same_opus_report_curves() -> None:
    percent_session = _opus_session("%", 1.0)
    ppm_session = _opus_session("ppm", 10_000.0)
    percent_controller = InterpretationCalculationController(percent_session)
    ppm_controller = InterpretationCalculationController(ppm_session)

    percent_controller.calculate_opus_curves()
    ppm_controller.calculate_opus_curves()

    for mnemonic in (
        "OPUS_TG_PCT",
        "OPUS3",
        "OPUS4",
        "OPUS_K1_3",
        "OPUS_1_5",
    ):
        percent_curve = percent_session.current_dataset.curve_by_mnemonic(mnemonic)
        ppm_curve = ppm_session.current_dataset.curve_by_mnemonic(mnemonic)
        np.testing.assert_allclose(percent_curve.values, ppm_curve.values, equal_nan=True)


def test_opus_report_is_marked_separate_and_uses_source_applicability_gates() -> None:
    session = _opus_session("%", 1.0)
    InterpretationCalculationController(session).calculate_opus_curves()

    report = build_opus_interpretation_report(session, threshold=3.0)

    assert report.report_profile == "opus"
    assert report.primary_mnemonic == "OPUS_TG_PCT"
    assert len(report.candidates) == 1
    assert "ОПУС" in fluid_hypothesis_label(report.candidates[0], AppLanguage.RU)
    assert any("OPUS interval means" in item for item in report.candidates[0].evidence)
    assert any("final automatic interpretation basis=" in item for item in report.candidates[0].evidence)
    assert any("отдельный дополнительный отчёт" in warning for warning in report.warnings)
    opus_method = next(method for method in report.methods if method.method.startswith("OPUS"))
    assert "OPUS3=p1×p2" in opus_method.calculation
    assert "Alekseev (2024)" in opus_method.source
    decision_method = next(
        method for method in report.methods if method.method.startswith("Whole-well")
    )
    assert "not a GOST/ISO" in decision_method.source
    html = hydrocarbon_interpretation_html(report, AppLanguage.RU)
    assert "Дополнительный отчёт ОПУС" in html
    assert "Расчёт и правило интерпретации" in html
    assert "OPUS3=p1×p2" in html


def test_opus_applicability_warning_does_not_delete_detected_gas_shows() -> None:
    session = _opus_session("%", 1.0)
    dataset = session.current_dataset
    assert dataset is not None
    peaks = {"C1": 0.50, "C2": 0.10, "C3": 0.06, "C4": 0.04, "C5": 0.03}
    for mnemonic, peak in peaks.items():
        curve = dataset.curve_by_mnemonic(mnemonic)
        assert curve is not None
        curve.values[:] = 0.0
        curve.values[40:43] = peak

    InterpretationCalculationController(session).calculate_opus_curves()
    report = build_opus_interpretation_report(session, threshold=3.0)

    assert report.baseline_median == 0.0
    assert len(report.candidates) == 1
    assert report.candidates[0].top_depth == 39.5
    assert report.candidates[0].bottom_depth == 42.5
    assert any("gas-show candidate retained" in item for item in report.candidates[0].evidence)
    assert any("аномалии сохранены" in warning for warning in report.warnings)


def test_opus_report_charts_prefer_opus_curves() -> None:
    session = _opus_session("%", 1.0)
    dataset = session.current_dataset
    assert dataset is not None
    InterpretationCalculationController(session).calculate_opus_curves()
    report = build_opus_interpretation_report(session)

    for selector in (screen_panel_curves, pdf_panel_curves):
        panels = dict(selector(report, dataset))
        assert panels["total"][0].metadata.original_mnemonic == "OPUS_TG_PCT"
        assert {curve.metadata.original_mnemonic for curve in panels["opus"]} == {
            "OPUS3",
            "OPUS4",
            "OPUS_K1_3",
            "OPUS_1_5",
        }


def test_opus_xlsx_contains_every_depth_row_and_working_percent_curves(tmp_path) -> None:
    session = _opus_session("ppm", 10_000.0)
    dataset = session.current_dataset
    assert dataset is not None
    InterpretationCalculationController(session).calculate_opus_curves()
    report = build_opus_interpretation_report(session)

    path = export_hydrocarbon_interpretation_xlsx(report, dataset, tmp_path / "opus.xlsx")
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        assert workbook["Интерпретация УВ"]["A1"].value.startswith("Дополнительный отчёт ОПУС")
        depth_sheet = workbook["Данные по глубине"]
        assert depth_sheet.max_row == dataset.depth.size + 1
        headers = [cell.value for cell in next(depth_sheet.iter_rows(min_row=1, max_row=1))]
        assert any(str(value).startswith("OPUS_C1_PCT ") for value in headers)
        assert any(str(value).startswith("OPUS_P1 ") for value in headers)
        assert any(str(value).startswith("OPUS3 ") for value in headers)
        assert not any(str(value).startswith("OPUS5_REF ") for value in headers)
        method_sheet = workbook["Методика"]
        method_headers = [cell.value for cell in next(method_sheet.iter_rows(max_row=1))]
        assert "Расчёт и правило интерпретации" in method_headers
        assert "Источник и степень подтверждения" in method_headers
        method_values = tuple(
            str(cell.value or "")
            for row in method_sheet.iter_rows(min_row=2, max_row=4)
            for cell in row
        )
        assert any("OPUS3=p1×p2" in value for value in method_values)
        assert any("Alekseev (2024)" in value for value in method_values)
    finally:
        workbook.close()
    docx_path = export_hydrocarbon_interpretation_docx(
        report, tmp_path / "opus.docx", dataset=dataset
    )
    with zipfile.ZipFile(docx_path) as package:
        document = package.read("word/document.xml").decode("utf-8")
        assert "Дополнительный отчёт ОПУС" in document
        assert "OPUS interval means" in document
        assert "Расчёт и правило интерпретации" in document
        assert "OPUS3=p1×p2" in document
