from __future__ import annotations

import csv
import os
import shutil
import subprocess
import tempfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from collections.abc import Iterator
from typing import Any

from geoworkbench.data.csv_adapter import CsvImportError, CsvImportPlan, CsvImportResult, import_csv


class ExcelImportError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class ExcelImportPlan:
    sheet_name: str
    header_row: int = 1
    index_column: str | None = None
    time_column: str | None = None
    date_format: str = "%Y-%m-%d"
    time_format: str = "%H:%M:%S"
    timezone: str | None = None

    def __post_init__(self) -> None:
        if not self.sheet_name.strip():
            raise ValueError("Нужно выбрать лист Excel")
        if self.header_row < 1:
            raise ValueError("Номер строки заголовка должен быть положительным")
        if self.time_column is not None and (not self.date_format or not self.time_format):
            raise ValueError("Для DATE+TIME нужны форматы обеих колонок")


@dataclass(frozen=True, slots=True)
class ExcelProbe:
    path: Path
    sheet_names: tuple[str, ...]
    columns: tuple[str, ...]
    preview_rows: tuple[tuple[str, ...], ...]


def probe_excel(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    header_row: int = 1,
    preview_rows: int = 20,
) -> ExcelProbe:
    source = _validate_source(path)
    with _readable_workbook(source, recalculate=False) as readable:
        workbook = _load_workbook(readable)
        try:
            sheets = tuple(workbook.sheetnames)
            selected = sheet_name or (sheets[0] if sheets else None)
            if selected is None or selected not in sheets:
                raise ExcelImportError(f"Лист Excel не найден: {selected or '—'}")
            rows = workbook[selected].iter_rows(min_row=header_row, values_only=True)
            header = next(rows, None)
            if header is None:
                raise ExcelImportError("Строка заголовка находится за пределами листа")
            columns = tuple(_cell_text(value).strip() for value in header)
            _validate_columns(columns)
            preview = tuple(
                tuple(_cell_text(value) for value in row[: len(columns)])
                for _, row in zip(range(preview_rows), rows)
            )
            return ExcelProbe(source, sheets, columns, preview)
        finally:
            workbook.close()


def excel_sheet_names(path: str | Path) -> tuple[str, ...]:
    source = _validate_source(path)
    with _readable_workbook(source, recalculate=False) as readable:
        workbook = _load_workbook(readable)
        try:
            return tuple(workbook.sheetnames)
        finally:
            workbook.close()


def import_excel(path: str | Path, plan: ExcelImportPlan) -> CsvImportResult:
    source = _validate_source(path)
    if plan.index_column is None:
        raise ExcelImportError("Нужно явно выбрать индексную колонку Excel")
    probe = probe_excel(source, sheet_name=plan.sheet_name, header_row=plan.header_row)
    temporary_name: str | None = None
    with _readable_workbook(source, recalculate=_contains_formulas(source)) as readable:
        workbook = _load_workbook(readable)
        try:
            rows = workbook[plan.sheet_name].iter_rows(min_row=plan.header_row, values_only=True)
            next(rows, None)
            with tempfile.NamedTemporaryFile(
                "w", encoding="utf-8", newline="", suffix=".csv", delete=False
            ) as temporary:
                temporary_name = temporary.name
                writer = csv.writer(temporary)
                writer.writerow(probe.columns)
                for row in rows:
                    values = tuple(row[: len(probe.columns)])
                    if all(value is None for value in values):
                        continue
                    writer.writerow(_cell_text(value) for value in values)
            result = import_csv(
                temporary_name,
                CsvImportPlan(
                    encoding="utf-8",
                    delimiter=",",
                    index_column=plan.index_column,
                    time_column=plan.time_column,
                    date_format=plan.date_format,
                    time_format=plan.time_format,
                    timezone=plan.timezone,
                ),
            )
            result.dataset.name = f"{source.stem} — {plan.sheet_name}"
            result.dataset.source_path = source
            return result
        except (CsvImportError, KeyError, ValueError) as exc:
            raise ExcelImportError(str(exc)) from exc
        finally:
            workbook.close()
            if temporary_name is not None:
                Path(temporary_name).unlink(missing_ok=True)


def _load_workbook(source: Path, *, data_only: bool = True) -> Any:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise ExcelImportError(
            "Компонент Excel не установлен; переустановите приложение: pip install -e ."
        ) from exc
    try:
        return load_workbook(source, read_only=True, data_only=data_only)
    except Exception as exc:
        raise ExcelImportError(f"Не удалось открыть Excel-файл: {source.name}") from exc


def _validate_source(path: str | Path) -> Path:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(source)
    if source.suffix.casefold() not in {".xls", ".xlsx", ".xlsm"}:
        raise ExcelImportError(f"Ожидался XLS/XLSX/XLSM-файл, получен: {source.suffix}")
    return source


def _contains_formulas(source: Path) -> bool:
    if source.suffix.casefold() == ".xls":
        return True
    workbook = _load_workbook(source, data_only=False)
    try:
        return any(
            cell.data_type == "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()


@contextmanager
def _readable_workbook(source: Path, *, recalculate: bool) -> Iterator[Path]:
    if source.suffix.casefold() != ".xls" and not recalculate:
        yield source
        return
    executable = _find_libreoffice()
    if executable is None:
        raise ExcelImportError(
            "Для XLS и пересчёта формул нужен LibreOffice Calc: "
            "sudo apt install libreoffice-calc"
        )
    with tempfile.TemporaryDirectory(prefix="geolog-excel-") as directory:
        root = Path(directory)
        output = root / "output"
        profile = root / "profile"
        output.mkdir()
        profile.mkdir()
        command = [
            executable,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            f"-env:UserInstallation={profile.as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output),
            str(source),
        ]
        environment = os.environ.copy()
        environment.update(
            HOME=str(root),
            XDG_CONFIG_HOME=str(root / "config"),
            XDG_CACHE_HOME=str(root / "cache"),
            XDG_RUNTIME_DIR=str(root / "runtime"),
        )
        for name in ("config", "cache", "runtime"):
            (root / name).mkdir()
        try:
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=120,
                env=environment,
            )
        except (OSError, subprocess.TimeoutExpired) as exc:
            raise ExcelImportError("LibreOffice не завершил подготовку книги за 120 секунд") from exc
        converted = output / f"{source.stem}.xlsx"
        if completed.returncode != 0 or not converted.exists():
            detail = completed.stderr.strip() or completed.stdout.strip()
            raise ExcelImportError(f"LibreOffice не смог подготовить книгу: {detail or source.name}")
        yield converted


def _find_libreoffice() -> str | None:
    configured = os.environ.get("GEOLOG_LIBREOFFICE")
    candidates = (
        configured,
        shutil.which("libreoffice"),
        shutil.which("soffice"),
        str(Path.cwd() / ".dev-tools/libreoffice/program/soffice"),
        "/usr/lib/libreoffice/program/soffice",
        "/snap/bin/libreoffice",
    )
    return next((candidate for candidate in candidates if candidate and Path(candidate).is_file()), None)


def _validate_columns(columns: tuple[str, ...]) -> None:
    if not columns or any(not value for value in columns):
        raise ExcelImportError("Заголовки Excel не должны быть пустыми")
    if len({value.casefold() for value in columns}) != len(columns):
        raise ExcelImportError("Заголовки Excel не должны повторяться")


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)
