from __future__ import annotations

import csv
import os
import tempfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from geoworkbench.data.number_format import format_decimal_number
from geoworkbench.domain.models import CurveData, Dataset, DatasetIndex, IndexRole, IndexType
from geoworkbench.services.dataset_selection import depth_interval_indices
from geoworkbench.services.las_parameter_resolver import LasParameterResolver, ParameterMatch
from geoworkbench.services.localization import AppLanguage
from geoworkbench.services.parameter_labels import localized_curve_name
from geoworkbench.services.text_normalization import clean_display_text, clean_mnemonic


class SelectionExportError(RuntimeError):
    pass


EXCEL_DECIMAL_NUMBER_FORMAT = "0.##############################"


_PARAMETER_NAME_OVERRIDES: dict[AppLanguage, dict[str, str]] = {
    AppLanguage.RU: {
        "C1": "Метан",
        "C2": "Этан",
        "C3": "Пропан",
        "C4": "Бутан",
        "C5": "Пентан",
        "IC4": "Изобутан",
        "NC4": "н-Бутан",
        "IC5": "Изопентан",
        "NC5": "н-Пентан",
        "TG": "Суммарный газ",
        "TOTAL_GAS": "Суммарный газ",
        "H2S": "Сероводород",
        "CO2": "Диоксид углерода",
    },
    AppLanguage.KK: {
        "C1": "Метан",
        "C2": "Этан",
        "C3": "Пропан",
        "C4": "Бутан",
        "C5": "Пентан",
        "IC4": "Изобутан",
        "NC4": "н-Бутан",
        "IC5": "Изопентан",
        "NC5": "н-Пентан",
        "TG": "Жалпы газ",
        "TOTAL_GAS": "Жалпы газ",
        "H2S": "Күкіртсутек",
        "CO2": "Көмірқышқыл газы",
    },
    AppLanguage.EN: {
        "C1": "Methane",
        "C2": "Ethane",
        "C3": "Propane",
        "C4": "Butane",
        "C5": "Pentane",
        "IC4": "Isobutane",
        "NC4": "n-Butane",
        "IC5": "Isopentane",
        "NC5": "n-Pentane",
        "TG": "Total gas",
        "TOTAL_GAS": "Total gas",
        "H2S": "Hydrogen sulfide",
        "CO2": "Carbon dioxide",
    },
}


@dataclass(frozen=True, slots=True)
class _ExportColumn:
    friendly_name: str
    technical_name: str
    unit: str
    canonical_mnemonic: str
    source_description: str
    recognized: bool
    confidence: float | None
    matched_by: str
    provenance: str

    @property
    def header(self) -> str:
        unit = self.unit or "—"
        return f"{self.friendly_name}\n{self.technical_name}\n[{unit}]"


def export_selection_text(
    dataset: Dataset,
    target: str | Path,
    curve_ids: list[str],
    depth_top: float,
    depth_bottom: float,
    *,
    delimiter: str = "\t",
    overwrite: bool = False,
) -> Path:
    if len(delimiter) != 1:
        raise ValueError("Разделитель должен состоять из одного символа")
    destination = Path(target)
    _validate_destination(destination, {".txt", ".csv"}, overwrite)
    indices, curves = _selection(dataset, curve_ids, depth_top, depth_bottom)
    temporary = _temporary_path(destination)
    try:
        with temporary.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream, delimiter=delimiter)
            writer.writerow(_headers(dataset, curves))
            for index in indices:
                writer.writerow(
                    [_number(dataset.depth[index])]
                    + [_number(curve.values[index]) for curve in curves]
                )
        os.replace(temporary, destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        raise SelectionExportError(f"Не удалось экспортировать интервал: {destination}") from exc
    return destination


def export_selection_excel(
    dataset: Dataset,
    target: str | Path,
    curve_ids: list[str],
    depth_top: float,
    depth_bottom: float,
    *,
    overwrite: bool = False,
    language: AppLanguage | str = AppLanguage.RU,
) -> Path:
    destination = Path(target)
    _validate_destination(destination, {".xlsx"}, overwrite)
    indices, curves = _selection(dataset, curve_ids, depth_top, depth_bottom)
    export_language = AppLanguage(language)
    export_indexes = _excel_indexes(dataset)
    index_columns = [
        _index_export_column(
            dataset,
            index,
            primary=index.index_id == dataset.active_index_id,
            language=export_language,
        )
        for index in export_indexes
    ]
    curve_columns = _curve_export_columns(curves, language=export_language)
    header: list[object] = [column.header for column in (*index_columns, *curve_columns)]
    rows: list[list[object]] = [header]
    rows.extend(
        [_excel_index_value(index_value, int(index)) for index_value in export_indexes]
        + [
            None if not np.isfinite(curve.values[index]) else float(curve.values[index])
            for curve in curves
        ]
        for index in indices
    )
    metadata_labels = _metadata_labels(export_language)
    metadata: list[list[object]] = [
        [metadata_labels["dataset"], clean_display_text(dataset.name)],
        [metadata_labels["depth_top"], depth_top],
        [metadata_labels["depth_bottom"], depth_bottom],
        [metadata_labels["rows"], int(indices.size)],
        [metadata_labels["source"], str(dataset.source_path or "")],
        [metadata_labels["language"], export_language.value],
    ]
    for index in export_indexes:
        mnemonic = clean_mnemonic(index.mnemonic)
        metadata.extend(
            [
                [f"{metadata_labels['index_type']} {mnemonic}", index.index_type.value],
                [f"{metadata_labels['timezone']} {mnemonic}", index.timezone or ""],
            ]
        )
    parameter_rows = _parameter_rows(
        (*index_columns, *curve_columns), language=export_language
    )
    temporary = _temporary_path(destination)
    try:
        _write_xlsx(temporary, rows, metadata, parameter_rows)
        os.replace(temporary, destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        raise SelectionExportError(f"Не удалось экспортировать Excel: {destination}") from exc
    return destination



def _curve_export_columns(
    curves: list[CurveData], *, language: AppLanguage
) -> list[_ExportColumn]:
    resolver = LasParameterResolver()
    result: list[_ExportColumn] = []
    for curve in curves:
        metadata = curve.metadata
        original = clean_mnemonic(metadata.original_mnemonic)
        description = clean_display_text(metadata.description or "")
        unit = clean_display_text(metadata.unit or "")
        matches = resolver.infer_curve(curve)
        match: ParameterMatch | None = matches[0] if matches else None
        canonical = clean_mnemonic(
            match.canonical_mnemonic
            if match is not None
            else metadata.canonical_mnemonic or original
        ).upper()
        friendly = _PARAMETER_NAME_OVERRIDES[language].get(canonical, "")
        if not friendly:
            friendly = localized_curve_name(
                canonical,
                description=description,
                unit=unit,
                language=language,
            ).strip()
        recognized = match is not None or bool(metadata.canonical_mnemonic)
        if not friendly or (
            not recognized
            and not description
            and friendly.casefold() == original.casefold()
        ):
            friendly = _unresolved_name(language)
        technical = original
        if canonical and canonical.casefold() != original.casefold():
            technical = f"{original} → {canonical}"
        result.append(
            _ExportColumn(
                friendly_name=friendly,
                technical_name=technical,
                unit=unit,
                canonical_mnemonic=canonical,
                source_description=description,
                recognized=recognized,
                confidence=match.confidence if match is not None else None,
                matched_by=match.matched_by if match is not None else "",
                provenance=clean_display_text(metadata.provenance),
            )
        )
    return result


def _index_export_column(
    dataset: Dataset,
    index: DatasetIndex,
    *,
    primary: bool,
    language: AppLanguage,
) -> _ExportColumn:
    mnemonic = clean_mnemonic(
        "DEPTH" if primary and index.role is IndexRole.DEPTH else index.mnemonic
    )
    unit = (
        "UTC"
        if index.index_type is IndexType.DATETIME and index.timezone
        else clean_display_text(index.unit or "")
    )
    friendly = _index_friendly_name(dataset, index, language=language)
    return _ExportColumn(
        friendly_name=friendly,
        technical_name=mnemonic,
        unit=unit,
        canonical_mnemonic=mnemonic,
        source_description="",
        recognized=True,
        confidence=1.0,
        matched_by="index",
        provenance="index",
    )


def _index_friendly_name(
    dataset: Dataset, index: DatasetIndex, *, language: AppLanguage
) -> str:
    if index.role is IndexRole.TIME:
        key = "datetime" if index.index_type is IndexType.DATETIME else "time"
    elif index.role is IndexRole.DEPTH:
        key = dataset.depth_domain.value
    else:
        key = "index"
    labels = {
        AppLanguage.RU: {
            "md": "Глубина по стволу",
            "tvd": "Истинная вертикальная глубина",
            "tvdss": "Вертикальная глубина относительно уровня моря",
            "time": "Время",
            "datetime": "Дата и время",
            "index": "Индекс",
        },
        AppLanguage.KK: {
            "md": "Оқпан бойынша тереңдік",
            "tvd": "Нақты тік тереңдік",
            "tvdss": "Теңіз деңгейіне қатысты тік тереңдік",
            "time": "Уақыт",
            "datetime": "Күні және уақыты",
            "index": "Индекс",
        },
        AppLanguage.EN: {
            "md": "Measured depth",
            "tvd": "True vertical depth",
            "tvdss": "True vertical depth subsea",
            "time": "Time",
            "datetime": "Date and time",
            "index": "Index",
        },
    }
    return labels[language].get(key, labels[language]["index"])


def _unresolved_name(language: AppLanguage) -> str:
    return {
        AppLanguage.RU: "Не распознано",
        AppLanguage.KK: "Анықталмаған параметр",
        AppLanguage.EN: "Unresolved parameter",
    }[language]


def _metadata_labels(language: AppLanguage) -> dict[str, str]:
    return {
        AppLanguage.RU: {
            "dataset": "Набор данных",
            "depth_top": "Верх интервала",
            "depth_bottom": "Низ интервала",
            "rows": "Количество строк",
            "source": "Источник",
            "language": "Язык экспорта",
            "index_type": "Тип индекса",
            "timezone": "Часовой пояс источника",
        },
        AppLanguage.KK: {
            "dataset": "Деректер жиыны",
            "depth_top": "Аралықтың төбесі",
            "depth_bottom": "Аралықтың табаны",
            "rows": "Жолдар саны",
            "source": "Дереккөз",
            "language": "Экспорт тілі",
            "index_type": "Индекс түрі",
            "timezone": "Дереккөздің уақыт белдеуі",
        },
        AppLanguage.EN: {
            "dataset": "Dataset",
            "depth_top": "Interval top",
            "depth_bottom": "Interval bottom",
            "rows": "Rows",
            "source": "Source",
            "language": "Export language",
            "index_type": "Index type",
            "timezone": "Source timezone",
        },
    }[language]


def _parameter_sheet_headers(language: AppLanguage) -> list[str]:
    return {
        AppLanguage.RU: [
            "№", "Понятное название", "Мнемоника LAS", "Каноническая мнемоника",
            "Единица", "Описание из LAS", "Распознано", "Уверенность",
            "Метод сопоставления", "Происхождение",
        ],
        AppLanguage.KK: [
            "№", "Түсінікті атауы", "LAS мнемоникасы", "Канондық мнемоника",
            "Өлшем бірлігі", "LAS сипаттамасы", "Анықталды", "Сенімділік",
            "Сәйкестендіру әдісі", "Шығу тегі",
        ],
        AppLanguage.EN: [
            "No.", "Readable name", "LAS mnemonic", "Canonical mnemonic",
            "Unit", "LAS description", "Resolved", "Confidence",
            "Match method", "Provenance",
        ],
    }[language]


def _parameter_rows(
    columns: tuple[_ExportColumn, ...], *, language: AppLanguage
) -> list[list[object]]:
    yes, no = {
        AppLanguage.RU: ("Да", "Нет"),
        AppLanguage.KK: ("Иә", "Жоқ"),
        AppLanguage.EN: ("Yes", "No"),
    }[language]
    rows: list[list[object]] = [_parameter_sheet_headers(language)]
    for number, column in enumerate(columns, start=1):
        original = column.technical_name.split(" → ", 1)[0]
        rows.append(
            [
                number,
                column.friendly_name,
                original,
                column.canonical_mnemonic,
                column.unit or "—",
                column.source_description or "—",
                yes if column.recognized else no,
                column.confidence,
                column.matched_by or "—",
                column.provenance or "—",
            ]
        )
    return rows

def _selection(
    dataset: Dataset, curve_ids: list[str], depth_top: float, depth_bottom: float
) -> tuple[np.ndarray, list[CurveData]]:
    if not curve_ids:
        raise ValueError("Выберите хотя бы один параметр")
    missing = [curve_id for curve_id in curve_ids if curve_id not in dataset.curves]
    if missing:
        raise KeyError(f"Кривые не найдены: {', '.join(missing)}")
    indices = depth_interval_indices(dataset, depth_top, depth_bottom)
    return indices, [dataset.curves[curve_id] for curve_id in curve_ids]


def _headers(dataset: Dataset, curves: list[CurveData]) -> list[object]:
    return [_index_header(dataset.active_index, primary=True), *_curve_headers(curves)]


def _curve_headers(curves: list[CurveData]) -> list[str]:
    return [
        f"{curve.metadata.original_mnemonic} [{curve.metadata.unit}]"
        if curve.metadata.unit
        else curve.metadata.original_mnemonic
        for curve in curves
    ]


def _index_header(index: DatasetIndex, *, primary: bool) -> str:
    qualifier = (
        "UTC"
        if index.index_type is IndexType.DATETIME and index.timezone is not None
        else index.unit or ""
    )
    mnemonic = "DEPTH" if primary and index.role is IndexRole.DEPTH else index.mnemonic
    return f"{mnemonic} [{qualifier}]" if qualifier else mnemonic


def _excel_indexes(dataset: Dataset) -> tuple[DatasetIndex, ...]:
    return (
        dataset.active_index,
        *(
            index
            for index in dataset.indexes.values()
            if index.index_id != dataset.active_index_id and index.role is IndexRole.TIME
        ),
    )


def _number(value: float) -> str:
    return "" if not np.isfinite(value) else format_decimal_number(float(value))


def _excel_index_value(index: DatasetIndex, row: int) -> object:
    if index.role is not IndexRole.TIME or index.index_type is not IndexType.DATETIME:
        value = np.asarray(index.values)[row]
        return None if not np.isfinite(value) else float(value)
    value = np.asarray(index.values)[row].astype("datetime64[ns]")
    if np.isnat(value):
        return None
    nanoseconds = int(value.astype(np.int64))
    seconds, remainder = divmod(nanoseconds, 1_000_000_000)
    return datetime(1970, 1, 1) + timedelta(seconds=seconds, microseconds=remainder // 1_000)


def _validate_destination(destination: Path, suffixes: set[str], overwrite: bool) -> None:
    if destination.suffix.casefold() not in suffixes:
        raise SelectionExportError("Неподдерживаемое расширение экспорта: " + destination.suffix)
    if destination.exists() and not overwrite:
        raise FileExistsError(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)


def _temporary_path(destination: Path) -> Path:
    descriptor, name = tempfile.mkstemp(
        prefix=f".{destination.name}.", suffix=".tmp", dir=destination.parent
    )
    os.close(descriptor)
    return Path(name)


def _write_xlsx(
    path: Path,
    data_rows: list[list[object]],
    metadata_rows: list[list[object]],
    parameter_rows: list[list[object]],
) -> None:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    for row in data_rows:
        data_sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    for cell in data_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    data_sheet.row_dimensions[1].height = 52

    for row in data_sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = EXCEL_DECIMAL_NUMBER_FORMAT
    data_sheet.freeze_panes = "A2"
    if data_rows and data_rows[0]:
        data_sheet.auto_filter.ref = f"A1:{get_column_letter(len(data_rows[0]))}{len(data_rows)}"
        for column_index in range(1, len(data_rows[0]) + 1):
            data_sheet.column_dimensions[get_column_letter(column_index)].width = (
                22 if column_index > 1 else 18
            )

    parameters_sheet = workbook.create_sheet("Parameters")
    for row in parameter_rows:
        parameters_sheet.append(row)
    for cell in parameters_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    parameters_sheet.freeze_panes = "A2"
    parameters_sheet.auto_filter.ref = f"A1:J{max(1, len(parameter_rows))}"
    parameter_widths = (6, 34, 20, 24, 15, 44, 14, 14, 22, 20)
    for index, width in enumerate(parameter_widths, start=1):
        parameters_sheet.column_dimensions[get_column_letter(index)].width = width
    for row in parameters_sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[5].alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(row[7].value, (int, float)):
            row[7].number_format = "0%"

    metadata_sheet = workbook.create_sheet("Metadata")
    for row in metadata_rows:
        metadata_sheet.append(row)
    for cell in metadata_sheet[1]:
        cell.font = Font(bold=True)
    metadata_sheet.column_dimensions["A"].width = 36
    metadata_sheet.column_dimensions["B"].width = 48
    metadata_sheet.freeze_panes = "A2"
    workbook.save(path)

