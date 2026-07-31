from __future__ import annotations

import os
from pathlib import Path
import tempfile

import numpy as np
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.comments import Comment  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from geoworkbench.data.hydrocarbon_interpretation_export import (
    HydrocarbonInterpretationExportError,
)
from geoworkbench.data.spreadsheet_safety import protect_spreadsheet_row
from geoworkbench.domain.models import Dataset
from geoworkbench.services.hydrocarbon_interpretation import (
    HydrocarbonCandidateInterval,
    HydrocarbonInterpretationReport,
    candidate_evidence_summary,
    fluid_hypothesis_basis,
    fluid_hypothesis_label,
)
from geoworkbench.services.interval_gas_statistics import (
    CandidateIntervalGasStatistics,
    IntervalCurveStatistics,
    build_candidate_interval_statistics,
    build_interval_statistics,
    enhanced_fluid_hypothesis_basis,
)
from geoworkbench.services.lba_standard import describe_lba_assessment
from geoworkbench.services.localization import AppLanguage


_EXCEL_MAX_ROWS = 1_048_576
_HEADERS = (
    "№",
    "Кровля",
    "Подошва",
    "Мощность",
    "Ед.",
    "Статус УВ-пласта",
    "Предварительная интерпретация",
    "Сила аномалии",
    "Исходный общий газ / единица",
    "Фон исходного газа",
    "Мин исходного газа",
    "Среднее исходного газа",
    "Медиана исходного газа",
    "Макс исходного газа",
    "Нормализованный газ / единица",
    "Фон нормализованного газа",
    "Мин нормализованного газа",
    "Медиана нормализованного газа",
    "Макс нормализованного газа",
    "Max robust z",
    "Точек выше порога",
    "C1-C5 в интервале",
    "Haworth / Pixler",
    "DEXP",
    "ЛБА и сопоставление",
    "Решение геолога / комментарий",
    "Основание",
)


def export_readable_hydrocarbon_interpretation_xlsx(
    report: HydrocarbonInterpretationReport,
    dataset: Dataset,
    target: str | Path,
    *,
    overwrite: bool = False,
) -> Path:
    if dataset.dataset_id != report.dataset_id:
        raise HydrocarbonInterpretationExportError(
            "Набор данных не соответствует сформированному отчёту интерпретации."
        )
    if dataset.depth.size + 1 > _EXCEL_MAX_ROWS:
        raise HydrocarbonInterpretationExportError(
            f"В наборе {dataset.depth.size} строк; лимит Excel — {_EXCEL_MAX_ROWS - 1}."
        )
    destination = Path(target)
    if destination.suffix.casefold() != ".xlsx":
        destination = destination.with_suffix(".xlsx")
    if destination.exists() and not overwrite:
        raise FileExistsError(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    try:
        main = workbook.active
        main.title = "Интерпретация УВ"
        _write_main_sheet(main, report, dataset)
        _write_methods_sheet(workbook, report)
        _write_whole_well_sheet(workbook, dataset)

        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{destination.stem}-",
            suffix=".xlsx",
            dir=destination.parent,
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        try:
            workbook.save(temporary)
            os.replace(temporary, destination)
        except Exception as exc:
            temporary.unlink(missing_ok=True)
            if isinstance(exc, (FileExistsError, HydrocarbonInterpretationExportError)):
                raise
            raise HydrocarbonInterpretationExportError(
                f"Не удалось экспортировать Excel: {destination}"
            ) from exc
    finally:
        workbook.close()
    return destination


def _write_main_sheet(sheet, report: HydrocarbonInterpretationReport, dataset: Dataset) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:AA1")
    sheet["A1"] = "Сводная интерпретация газового каротажа и УВ-интервалов"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17365D")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    metadata = (
        ("Проект", report.project_name, "Скважина", report.well_name),
        ("Набор данных", report.dataset_name, "Сформирован", report.generated_at),
        (
            "Основная газовая кривая",
            report.primary_mnemonic or "-",
            "Порог robust z",
            report.threshold,
        ),
        (
            "Кандидатных УВ-интервалов",
            len(report.candidates),
            "Подтверждено геологом",
            len(report.manual_intervals),
        ),
    )
    for row_index, row in enumerate(metadata, start=2):
        sheet.cell(row_index, 1, row[0])
        sheet.cell(row_index, 2, row[1])
        sheet.cell(row_index, 5, row[2])
        sheet.cell(row_index, 6, row[3])
        for column in (1, 5):
            sheet.cell(row_index, column).font = Font(bold=True, color="17365D")
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=4)
        sheet.merge_cells(start_row=row_index, start_column=6, end_row=row_index, end_column=9)

    sheet.merge_cells("A7:AA7")
    sheet["A7"] = (
        "Примечание: 0 — реальное нулевое измерение. Пустая ячейка означает, что кривая "
        "или корректные отсчёты отсутствуют. Фон указан в единицах соответствующей газовой "
        "кривой."
    )
    sheet["A7"].fill = PatternFill("solid", fgColor="FFF2CC")
    sheet["A7"].font = Font(italic=True, color="7F6000")
    sheet["A7"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[7].height = 34

    group_specs = (
        ("A8:H8", "Интервал и интерпретация", "D9EAF7"),
        ("I8:N8", "Исходный общий газ", "E2F0D9"),
        ("O8:S8", "Нормализованный газ", "FCE4D6"),
        ("T8:U8", "Параметры аномалии", "E4DFEC"),
        ("V8:AA8", "Контекст и геологический контроль", "DDEBF7"),
    )
    for range_name, title, color in group_specs:
        sheet.merge_cells(range_name)
        cell = sheet[range_name.split(":", 1)[0]]
        cell.value = title
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="17365D")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.append([])
    for column, value in enumerate(_HEADERS, start=1):
        cell = sheet.cell(9, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="315A7D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[9].height = 48

    rows: list[tuple[object, ...]] = []
    candidate_ranges: list[tuple[float, float]] = []
    for index, candidate in enumerate(report.candidates, start=1):
        statistics = build_candidate_interval_statistics(dataset, candidate)
        matching_manual = _matching_manual_intervals(
            report,
            candidate.top_depth,
            candidate.bottom_depth,
        )
        rows.append(
            _candidate_row(
                index,
                report,
                candidate,
                statistics,
                matching_manual,
            )
        )
        candidate_ranges.append((candidate.top_depth, candidate.bottom_depth))

    manual_only = [
        item
        for item in report.manual_intervals
        if not any(
            _overlaps(item.top_depth, item.bottom_depth, top, bottom)
            for top, bottom in candidate_ranges
        )
    ]
    primary_mnemonic = _first_primary_name(report.primary_mnemonic)
    for manual_index, item in enumerate(manual_only, start=1):
        statistics = build_interval_statistics(
            dataset,
            item.top_depth,
            item.bottom_depth,
            primary_mnemonic=primary_mnemonic,
        )
        rows.append(_manual_row(manual_index, report, item, statistics))

    if not rows:
        rows.append(
            (
                1,
                None,
                None,
                None,
                report.depth_unit,
                "Кандидатные УВ-интервалы не найдены",
                "Проверьте порог robust z и доступность газовых кривых",
                None,
                *(None for _ in range(len(_HEADERS) - 8)),
            )
        )

    start_row = 10
    for row in rows:
        sheet.append(protect_spreadsheet_row(row))
    last_row = start_row + len(rows) - 1

    table = Table(displayName="HydrocarbonIntervals", ref=f"A9:AA{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "I10"
    sheet.auto_filter.ref = f"A9:AA{last_row}"

    thin = Side(style="thin", color="B7C9D6")
    for row in sheet.iter_rows(min_row=10, max_row=last_row, min_col=1, max_col=27):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for row_index in range(10, last_row + 1):
        sheet.row_dimensions[row_index].height = 58
        strength = sheet.cell(row_index, 8).value
        if strength == "высокая":
            sheet.cell(row_index, 8).fill = PatternFill("solid", fgColor="F4CCCC")
        elif strength == "средняя":
            sheet.cell(row_index, 8).fill = PatternFill("solid", fgColor="FCE5CD")
        elif strength == "низкая":
            sheet.cell(row_index, 8).fill = PatternFill("solid", fgColor="FFF2CC")
        if "Подтвержден" in str(sheet.cell(row_index, 6).value or ""):
            sheet.cell(row_index, 6).fill = PatternFill("solid", fgColor="D9EAD3")

    numeric_columns = (2, 3, 4, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20)
    for column in numeric_columns:
        for row_index in range(10, last_row + 1):
            sheet.cell(row_index, column).number_format = "0.######"
    sheet["J9"].comment = Comment(
        "Фон — robust-медиана log1p-преобразованной газовой кривой, возвращённая в "
        "исходные единицы.",
        "GEOLOG GASRATIO@Pixler",
    )
    sheet["P9"].comment = Comment(
        "Фон рассчитан отдельно для выбранной нормализованной газовой кривой.",
        "GEOLOG GASRATIO@Pixler",
    )

    widths = (
        6,
        12,
        12,
        11,
        8,
        24,
        38,
        15,
        24,
        15,
        15,
        17,
        17,
        15,
        26,
        16,
        16,
        17,
        16,
        14,
        15,
        38,
        36,
        24,
        42,
        42,
        56,
    )
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = f"A9:AA{last_row}"
    sheet.print_title_rows = "1:9"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _candidate_row(
    index: int,
    report: HydrocarbonInterpretationReport,
    candidate: HydrocarbonCandidateInterval,
    statistics: CandidateIntervalGasStatistics,
    manual_intervals: tuple[object, ...],
) -> tuple[object, ...]:
    raw = statistics.raw_total
    normalized = statistics.primary
    status = "Кандидат УВ-пласта"
    geologist = "Требуется подтверждение геологом"
    if manual_intervals:
        status = "Подтвержден геологом"
        geologist = " | ".join(
            f"{item.interpretation_name}: {item.label or item.interval_type}; "
            f"{item.comment}".strip("; ")
            for item in manual_intervals
        )
    base = fluid_hypothesis_basis(candidate, AppLanguage.RU)
    readable_basis = enhanced_fluid_hypothesis_basis(
        base,
        candidate,
        statistics,
        AppLanguage.RU,
    )
    return (
        index,
        candidate.top_depth,
        candidate.bottom_depth,
        candidate.bottom_depth - candidate.top_depth,
        report.depth_unit,
        status,
        fluid_hypothesis_label(candidate, AppLanguage.RU),
        {"low": "низкая", "medium": "средняя", "high": "высокая"}.get(
            candidate.anomaly_strength,
            candidate.anomaly_strength,
        ),
        _curve_identity(raw),
        _stat(raw, "background"),
        _stat(raw, "minimum"),
        _stat(raw, "mean"),
        _stat(raw, "median"),
        _stat(raw, "maximum"),
        _curve_identity(normalized),
        _stat(normalized, "background"),
        _stat(normalized, "minimum"),
        _stat(normalized, "median"),
        _stat(normalized, "maximum"),
        candidate.max_robust_z,
        candidate.sample_count,
        _components_text(statistics.components),
        _haworth_pixler_text(candidate),
        _dexp_text(statistics.dexp),
        _lba_text(candidate),
        geologist,
        readable_basis or candidate_evidence_summary(candidate),
    )


def _manual_row(index: int, report, item, statistics) -> tuple[object, ...]:
    raw = statistics.raw_total
    normalized = statistics.primary
    return (
        f"Г-{index}",
        item.top_depth,
        item.bottom_depth,
        item.bottom_depth - item.top_depth,
        report.depth_unit,
        "Подтвержден геологом",
        item.interpretation_name,
        None,
        _curve_identity(raw),
        _stat(raw, "background"),
        _stat(raw, "minimum"),
        _stat(raw, "mean"),
        _stat(raw, "median"),
        _stat(raw, "maximum"),
        _curve_identity(normalized),
        _stat(normalized, "background"),
        _stat(normalized, "minimum"),
        _stat(normalized, "median"),
        _stat(normalized, "maximum"),
        None,
        None,
        _components_text(statistics.components),
        None,
        _dexp_text(statistics.dexp),
        None,
        f"{item.label or item.interval_type}; {item.comment}".strip("; "),
        "Интервал внесён и подтверждён геологом.",
    )


def _write_methods_sheet(workbook: Workbook, report: HydrocarbonInterpretationReport) -> None:
    sheet = workbook.create_sheet("Методика")
    sheet.sheet_view.showGridLines = False
    sheet.append(
        protect_spreadsheet_row(
            ("Метод", "Ожидаемые кривые", "Доступные кривые", "Статус", "Источник")
        )
    )
    for method in report.methods:
        sheet.append(
            protect_spreadsheet_row(
                (
                    method.method,
                    ", ".join(method.curve_mnemonics),
                    ", ".join(method.available_mnemonics),
                    "доступен" if method.available else "нет данных",
                    method.source,
                )
            )
        )
    sheet.append(())
    sheet.append(protect_spreadsheet_row(("Ограничения методики",)))
    for warning in report.warnings:
        sheet.append(protect_spreadsheet_row((warning,)))
    _format_auxiliary_sheet(sheet, widths=(38, 46, 46, 16, 90))


def _write_whole_well_sheet(workbook: Workbook, dataset: Dataset) -> None:
    sheet = workbook.create_sheet("Данные по глубине")
    curves = tuple(dataset.curves.values())
    index_header = (
        f"{dataset.active_index.mnemonic} [{dataset.active_index.unit}]"
        if dataset.active_index.unit
        else dataset.active_index.mnemonic
    )
    curve_headers = tuple(
        f"{curve.metadata.original_mnemonic} [{curve.metadata.unit}]"
        if curve.metadata.unit
        else curve.metadata.original_mnemonic
        for curve in curves
    )
    sheet.append(protect_spreadsheet_row((index_header, *curve_headers)))
    index_values = np.asarray(dataset.active_index.values)
    for row_index in range(index_values.size):
        row: list[object] = [_excel_value(index_values[row_index])]
        row.extend(_excel_value(curve.values[row_index]) for curve in curves)
        sheet.append(protect_spreadsheet_row(row))
    _format_auxiliary_sheet(
        sheet,
        widths=(18, *(14 for _curve in curves)),
        wrap_header=True,
    )
    sheet.sheet_state = "hidden"


def _format_auxiliary_sheet(
    sheet,
    *,
    widths: tuple[int, ...],
    wrap_header: bool = False,
) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="315A7D")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=wrap_header,
        )
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _matching_manual_intervals(report, top: float, bottom: float) -> tuple[object, ...]:
    return tuple(
        item
        for item in report.manual_intervals
        if _overlaps(top, bottom, item.top_depth, item.bottom_depth)
    )


def _overlaps(
    left_top: float,
    left_bottom: float,
    right_top: float,
    right_bottom: float,
) -> bool:
    return max(left_top, right_top) < min(left_bottom, right_bottom)


def _curve_identity(item: IntervalCurveStatistics | None) -> str | None:
    if item is None:
        return None
    return f"{item.mnemonic} [{item.unit}]" if item.unit else item.mnemonic


def _stat(item: IntervalCurveStatistics | None, field: str) -> float | None:
    if item is None or not item.has_values:
        return None
    return getattr(item, field)


def _components_text(components: tuple[IntervalCurveStatistics, ...]) -> str:
    if not components:
        return "нет данных"
    families = {_component_family(item.mnemonic): item for item in components}
    heavier = [families.get(name) for name in ("C2", "C3", "C4", "C5")]
    heavier_present = [item for item in heavier if item is not None]
    parts: list[str] = []
    c1 = families.get("C1")
    if c1 is not None:
        parts.append(f"C1: {_range(c1)} {c1.unit}".strip())
    if heavier_present and all(item.is_zero_only for item in heavier_present):
        parts.append("C2-C5: выше нуля не зарегистрированы")
    else:
        parts.extend(
            f"{name}: {_range(item)} {item.unit}".strip()
            for name, item in (
                (name, families.get(name))
                for name in ("C2", "C3", "C4", "C5")
            )
            if item is not None
        )
    return "; ".join(parts)


def _component_family(mnemonic: str) -> str:
    upper = mnemonic.upper()
    for name in ("C1", "C2", "C3", "C4", "C5"):
        if name in upper:
            return name
    return mnemonic


def _range(item: IntervalCurveStatistics) -> str:
    if item.minimum is None or item.maximum is None:
        return "нет данных"
    return f"{item.minimum:.6g}-{item.maximum:.6g}"


def _haworth_pixler_text(candidate: HydrocarbonCandidateInterval) -> str:
    parts = [
        f"Wh={_optional(candidate.interval_wetness)}",
        f"Bh={_optional(candidate.interval_balance)}",
        f"Ch={_optional(candidate.interval_character)}",
    ]
    if candidate.pixler_assessment is not None:
        pixler = candidate.pixler_assessment
        parts.append(
            f"Pixler={pixler.code}; C1/C2={pixler.c1_c2:.6g}; "
            f"профиль={pixler.profile_shape or 'недостаточно данных'}"
        )
    return "; ".join(parts)


def _dexp_text(item: IntervalCurveStatistics | None) -> str | None:
    if item is None or not item.has_values:
        return None
    return (
        f"{item.mnemonic}: min {_optional(item.minimum)}; "
        f"медиана {_optional(item.median)}; max {_optional(item.maximum)}"
    )


def _lba_text(candidate: HydrocarbonCandidateInterval) -> str:
    descriptions = [
        describe_lba_assessment(item, AppLanguage.RU)
        for item in candidate.lba_assessments
    ]
    correlation = {
        "gas_only": "только газовые данные",
        "concordant": "признаки согласуются",
        "partial": "частичное согласие",
        "divergent": "признаки расходятся",
        "mixed": "смешанное сопоставление",
        "indeterminate": "недостаточно данных",
    }.get(candidate.gas_lba_correlation, candidate.gas_lba_correlation)
    if descriptions:
        return "; ".join((*descriptions, f"Сопоставление: {correlation}"))
    return f"ЛБА отсутствует; сопоставление: {correlation}"


def _optional(value: float | None) -> str:
    return "нет данных" if value is None or not np.isfinite(value) else f"{value:.6g}"


def _first_primary_name(primary: str | None) -> str | None:
    if not primary:
        return None
    name = primary.split(" | ", 1)[0].strip()
    for prefix in ("server: ", "local-calculation: "):
        if name.startswith(prefix):
            return name[len(prefix) :].strip()
    return name


def _excel_value(value: object) -> object:
    if isinstance(value, np.datetime64):
        return str(value)
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and not np.isfinite(value):
        return None
    return value


__all__ = ["export_readable_hydrocarbon_interpretation_xlsx"]
