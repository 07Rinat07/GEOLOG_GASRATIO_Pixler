from __future__ import annotations

import os
from pathlib import Path
import tempfile
import zipfile
from xml.sax.saxutils import escape as xml_escape

import numpy as np
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from geoworkbench.data.spreadsheet_safety import protect_spreadsheet_row
from geoworkbench.domain.models import Dataset
from geoworkbench.services.hydrocarbon_interpretation import (
    HydrocarbonInterpretationReport,
    candidate_evidence_summary,
    fluid_hypothesis_basis,
    fluid_hypothesis_label,
)
from geoworkbench.services.lba_standard import describe_lba_assessment
from geoworkbench.services.localization import AppLanguage


class HydrocarbonInterpretationExportError(RuntimeError):
    pass


_EXCEL_MAX_ROWS = 1_048_576


def export_hydrocarbon_interpretation_xlsx(
    report: HydrocarbonInterpretationReport,
    dataset: Dataset,
    target: str | Path,
    *,
    overwrite: bool = False,
) -> Path:
    _validate_dataset(report, dataset)
    destination = _prepare_target(target, ".xlsx", overwrite=overwrite)
    if dataset.depth.size + 1 > _EXCEL_MAX_ROWS:
        raise HydrocarbonInterpretationExportError(
            f"В наборе {dataset.depth.size} строк; лимит листа Excel — {_EXCEL_MAX_ROWS - 1}."
        )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary_rows = (
        ("Project", report.project_name),
        ("Well", report.well_name),
        ("Dataset", report.dataset_name),
        ("Generated", report.generated_at),
        ("Primary gas curve", report.primary_mnemonic or ""),
        ("Robust z threshold", report.threshold),
        ("Candidate intervals", len(report.candidates)),
        ("Geologist-confirmed intervals", len(report.manual_intervals)),
        ("Status", "Candidates require geologist confirmation"),
    )
    for summary_row in summary_rows:
        summary.append(protect_spreadsheet_row(summary_row))
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 80
    summary["A1"].font = Font(bold=True)

    candidates = workbook.create_sheet("Candidate intervals")
    candidate_headers = (
        "top_depth",
        "bottom_depth",
        "depth_unit",
        "relative_anomaly_strength",
        "flagged_samples",
        "primary_curve",
        "max_robust_z",
        "max_primary_value",
        "preliminary_fluid_interpretation",
        "interval_wetness_pct",
        "background_wetness_pct",
        "wetness_relative_robust_z",
        "interval_balance_bh",
        "interval_character_ch",
        "pixler_interpretation",
        "pixler_c1_c2",
        "pixler_profile_shape",
        "pixler_possible_water_association",
        "lba_standard_assessments",
        "gas_lba_correlation",
        "context_medians",
        "evidence",
        "review_status",
    )
    candidates.append(protect_spreadsheet_row(candidate_headers))
    for candidate in report.candidates:
        candidates.append(
            protect_spreadsheet_row(
                (
                    candidate.top_depth,
                    candidate.bottom_depth,
                    report.depth_unit,
                    candidate.anomaly_strength,
                    candidate.sample_count,
                    candidate.primary_mnemonic,
                    candidate.max_robust_z,
                    candidate.max_primary_value,
                    candidate.fluid_hypothesis,
                    candidate.interval_wetness,
                    candidate.background_wetness,
                    candidate.wetness_robust_z,
                    candidate.interval_balance,
                    candidate.interval_character,
                    candidate.pixler_assessment.code
                    if candidate.pixler_assessment is not None
                    else "",
                    candidate.pixler_assessment.c1_c2
                    if candidate.pixler_assessment is not None
                    else None,
                    candidate.pixler_assessment.profile_shape
                    if candidate.pixler_assessment is not None
                    else "",
                    (
                        "yes"
                        if candidate.pixler_assessment.water_association_possible
                        else "no"
                    )
                    if candidate.pixler_assessment is not None
                    else "",
                    "; ".join(
                        describe_lba_assessment(assessment, AppLanguage.RU)
                        for assessment in candidate.lba_assessments
                    ),
                    candidate.gas_lba_correlation,
                    "; ".join(f"{name}={value:.6g}" for name, value in candidate.metrics),
                    "; ".join(candidate.evidence),
                    "candidate — geologist confirmation required",
                )
            )
        )
    _format_table_sheet(
        candidates,
        widths=(
            14,
            14,
            12,
            20,
            16,
            18,
            16,
            18,
            38,
            20,
            20,
            24,
            18,
            18,
            38,
            18,
            18,
            24,
            70,
            24,
            48,
            72,
            38,
        ),
    )

    manual = workbook.create_sheet("Manual intervals")
    manual_headers = (
        "interpretation",
        "top_depth",
        "bottom_depth",
        "depth_unit",
        "interval_type",
        "label",
        "comment",
    )
    manual.append(protect_spreadsheet_row(manual_headers))
    for manual_interval in report.manual_intervals:
        manual.append(
            protect_spreadsheet_row(
                (
                    manual_interval.interpretation_name,
                    manual_interval.top_depth,
                    manual_interval.bottom_depth,
                    report.depth_unit,
                    manual_interval.interval_type,
                    manual_interval.label,
                    manual_interval.comment,
                )
            )
        )
    _format_table_sheet(manual, widths=(28, 14, 14, 12, 22, 38, 70))

    methods = workbook.create_sheet("Methods")
    methods.append(
        protect_spreadsheet_row(
            ("method", "expected_curves", "available_curves", "available", "source")
        )
    )
    for method in report.methods:
        methods.append(
            protect_spreadsheet_row(
                (
                    method.method,
                    ", ".join(method.curve_mnemonics),
                    ", ".join(method.available_mnemonics),
                    "yes" if method.available else "no",
                    method.source,
                )
            )
        )
    methods.append(())
    methods.append(protect_spreadsheet_row(("Method limitations",)))
    for warning in report.warnings:
        methods.append(protect_spreadsheet_row((warning,)))
    _format_table_sheet(methods, widths=(38, 42, 42, 14, 90))

    whole_well = workbook.create_sheet("Whole well")
    curves = tuple(dataset.curves.values())
    index_header = (
        f"{dataset.active_index.mnemonic} [{dataset.active_index.unit}]"
        if dataset.active_index.unit
        else dataset.active_index.mnemonic
    )
    curve_headers = tuple(
        f"{curve.metadata.original_mnemonic} [{curve.metadata.unit}]"
        if curve.metadata.unit
        else curve.metadata.original_mnemonic
        for curve in curves
    )
    whole_well.append(protect_spreadsheet_row((index_header, *curve_headers)))
    index_values = np.asarray(dataset.active_index.values)
    for row_index in range(index_values.size):
        whole_row: list[object] = [_excel_value(index_values[row_index])]
        whole_row.extend(_excel_value(curve.values[row_index]) for curve in curves)
        whole_well.append(protect_spreadsheet_row(whole_row))
    _format_table_sheet(
        whole_well,
        widths=(18, *(14 for _curve in curves)),
        wrap_header=True,
    )

    temporary = _temporary_path(destination)
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        if isinstance(exc, (FileExistsError, HydrocarbonInterpretationExportError)):
            raise
        raise HydrocarbonInterpretationExportError(
            f"Не удалось экспортировать Excel: {destination}"
        ) from exc
    finally:
        workbook.close()
    return destination


def export_hydrocarbon_interpretation_docx(
    report: HydrocarbonInterpretationReport,
    target: str | Path,
    *,
    overwrite: bool = False,
) -> Path:
    destination = _prepare_target(target, ".docx", overwrite=overwrite)
    temporary = _temporary_path(destination)
    try:
        _write_docx(temporary, report)
        os.replace(temporary, destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        if isinstance(exc, (FileExistsError, HydrocarbonInterpretationExportError)):
            raise
        raise HydrocarbonInterpretationExportError(
            f"Не удалось экспортировать Word: {destination}"
        ) from exc
    return destination


def _format_table_sheet(sheet, *, widths: tuple[int, ...], wrap_header: bool = False) -> None:
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="315A7D")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=wrap_header,
            )
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _excel_value(value: object) -> object:
    if isinstance(value, np.datetime64):
        return str(value)
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and not np.isfinite(value):
        return None
    return value


def _write_docx(path: Path, report: HydrocarbonInterpretationReport) -> None:
    body: list[str] = [
        _paragraph("Отчёт по интерпретации газового каротажа", style="Title"),
        _paragraph(f"Проект: {report.project_name}"),
        _paragraph(f"Скважина: {report.well_name}"),
        _paragraph(f"Набор данных: {report.dataset_name}"),
        _paragraph(f"Сформирован: {report.generated_at}"),
        _paragraph(f"Основная газовая кривая: {report.primary_mnemonic or '—'}"),
        _paragraph(f"Порог robust z: {report.threshold:.2f}"),
        _paragraph("Методы и доступность", style="Heading1"),
        _table(
            ("Метод", "Доступные кривые", "Источник"),
            tuple(
                (
                    method.method,
                    ", ".join(method.available_mnemonics) or "нет",
                    method.source,
                )
                for method in report.methods
            ),
            widths=(4_600, 3_000, 7_500),
        ),
        _paragraph("Кандидатные интервалы УВ-проявлений", style="Heading1"),
    ]
    if report.candidates:
        body.append(
            _table(
                (
                    "Интервал",
                    "Сила аномалии",
                    "Предварительная интерпретация",
                    "Точек выше порога",
                    "Основание",
                ),
                tuple(
                    (
                        f"{item.top_depth:.2f}–{item.bottom_depth:.2f} {report.depth_unit}",
                        {
                            "low": "низкая",
                            "medium": "средняя",
                            "high": "высокая",
                        }[item.anomaly_strength],
                        fluid_hypothesis_label(item, AppLanguage.RU),
                        str(item.sample_count),
                        candidate_evidence_summary(item),
                    )
                    for item in report.candidates
                ),
                widths=(2_200, 2_000, 3_700, 1_800, 5_400),
            )
        )
    else:
        body.append(_paragraph("Кандидатные интервалы по выбранному порогу не найдены."))
    body.append(_paragraph("Сопоставление методов по интервалам", style="Heading1"))
    if report.candidates:
        for item in report.candidates:
            body.append(
                _paragraph(
                    f"{item.top_depth:.2f}–{item.bottom_depth:.2f} {report.depth_unit}: "
                    f"{fluid_hypothesis_label(item, AppLanguage.RU)}. "
                    f"{fluid_hypothesis_basis(item, AppLanguage.RU)}"
                )
            )
    else:
        body.append(_paragraph("Кандидатные интервалы по выбранному порогу не найдены."))
    body.append(_paragraph("Интервалы, подтверждённые геологом", style="Heading1"))
    if report.manual_intervals:
        body.append(
            _table(
                ("Интерпретация", "Интервал", "Тип", "Подпись", "Комментарий"),
                tuple(
                    (
                        item.interpretation_name,
                        f"{item.top_depth:.2f}–{item.bottom_depth:.2f} {report.depth_unit}",
                        item.interval_type,
                        item.label,
                        item.comment,
                    )
                    for item in report.manual_intervals
                ),
                widths=(2_600, 2_200, 2_200, 4_000, 4_100),
            )
        )
    else:
        body.append(_paragraph("Подтверждённые геологом интервалы пока не заполнены."))
    body.append(_paragraph("Ограничения методики", style="Heading1"))
    body.extend(_paragraph(f"• {warning}") for warning in report.warnings)
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + "".join(body)
        + '<w:sectPr><w:pgSz w:w="16838" w:h="11906" w:orient="landscape"/>'
        '<w:pgMar w:top="850" w:right="850" w:bottom="850" w:left="850" '
        'w:header="708" w:footer="708" w:gutter="0"/></w:sectPr>'
        "</w:body></w:document>"
    )
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            '<Override PartName="/word/styles.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
            "</Types>",
        )
        package.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="word/document.xml"/></Relationships>',
        )
        package.writestr("word/document.xml", document_xml)
        package.writestr(
            "word/_rels/document.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
        )
        package.writestr("word/styles.xml", _docx_styles())


def _paragraph(text: str, *, style: str | None = None) -> str:
    style_xml = f'<w:pPr><w:pStyle w:val="{style}"/></w:pPr>' if style else ""
    return f"<w:p>{style_xml}<w:r><w:t xml:space=\"preserve\">{_xml_text(text)}</w:t></w:r></w:p>"


def _table(
    headers: tuple[str, ...],
    rows: tuple[tuple[str, ...], ...],
    *,
    widths: tuple[int, ...],
) -> str:
    if len(widths) != len(headers) or any(len(row) != len(headers) for row in rows):
        raise ValueError("Геометрия таблицы отчёта не соответствует числу колонок")
    grid = "".join(f'<w:gridCol w:w="{width}"/>' for width in widths)
    header = _table_row(headers, widths, header=True)
    body = "".join(_table_row(row, widths) for row in rows)
    return (
        '<w:tbl><w:tblPr><w:tblW w:w="15100" w:type="dxa"/>'
        '<w:tblLayout w:type="fixed"/><w:tblCellMar>'
        '<w:top w:w="90" w:type="dxa"/><w:left w:w="90" w:type="dxa"/>'
        '<w:bottom w:w="90" w:type="dxa"/><w:right w:w="90" w:type="dxa"/>'
        '</w:tblCellMar><w:tblBorders>'
        '<w:top w:val="single" w:sz="6" w:color="8290A3"/>'
        '<w:left w:val="single" w:sz="6" w:color="8290A3"/>'
        '<w:bottom w:val="single" w:sz="6" w:color="8290A3"/>'
        '<w:right w:val="single" w:sz="6" w:color="8290A3"/>'
        '<w:insideH w:val="single" w:sz="4" w:color="8290A3"/>'
        '<w:insideV w:val="single" w:sz="4" w:color="8290A3"/>'
        "</w:tblBorders></w:tblPr>"
        f"<w:tblGrid>{grid}</w:tblGrid>{header}{body}</w:tbl>"
    )


def _table_row(
    values: tuple[str, ...],
    widths: tuple[int, ...],
    *,
    header: bool = False,
) -> str:
    cells = []
    for value, width in zip(values, widths, strict=True):
        run_properties = (
            "<w:rPr><w:b/><w:sz w:val=\"18\"/></w:rPr>"
            if header
            else '<w:rPr><w:sz w:val="18"/></w:rPr>'
        )
        shading = '<w:shd w:val="clear" w:fill="DCE8F4"/>' if header else ""
        cells.append(
            f'<w:tc><w:tcPr><w:tcW w:w="{width}" w:type="dxa"/>'
            f'<w:vAlign w:val="center"/>{shading}</w:tcPr>'
            '<w:p><w:pPr><w:spacing w:after="0"/></w:pPr>'
            f'<w:r>{run_properties}<w:t xml:space="preserve">'
            f"{_xml_text(value)}</w:t></w:r></w:p></w:tc>"
        )
    row_properties = (
        "<w:trPr><w:tblHeader/><w:cantSplit/></w:trPr>"
        if header
        else "<w:trPr><w:cantSplit/></w:trPr>"
    )
    return f"<w:tr>{row_properties}" + "".join(cells) + "</w:tr>"


def _xml_text(value: object) -> str:
    text = "".join(
        character
        for character in str(value)
        if character in "\t\n\r" or ord(character) >= 0x20
    )
    return xml_escape(text)


def _docx_styles() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:style w:type="paragraph" w:default="1" w:styleId="Normal">'
        '<w:name w:val="Normal"/><w:rPr><w:sz w:val="20"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Title">'
        '<w:name w:val="Title"/><w:basedOn w:val="Normal"/>'
        '<w:rPr><w:b/><w:sz w:val="34"/></w:rPr></w:style>'
        '<w:style w:type="paragraph" w:styleId="Heading1">'
        '<w:name w:val="heading 1"/><w:basedOn w:val="Normal"/>'
        '<w:rPr><w:b/><w:sz w:val="26"/></w:rPr></w:style>'
        '<w:style w:type="table" w:styleId="TableGrid"><w:name w:val="Table Grid"/>'
        '<w:tblPr><w:tblBorders>'
        '<w:top w:val="single" w:sz="4" w:color="808080"/>'
        '<w:left w:val="single" w:sz="4" w:color="808080"/>'
        '<w:bottom w:val="single" w:sz="4" w:color="808080"/>'
        '<w:right w:val="single" w:sz="4" w:color="808080"/>'
        '<w:insideH w:val="single" w:sz="4" w:color="808080"/>'
        '<w:insideV w:val="single" w:sz="4" w:color="808080"/>'
        "</w:tblBorders></w:tblPr></w:style></w:styles>"
    )


def _validate_dataset(report: HydrocarbonInterpretationReport, dataset: Dataset) -> None:
    if report.dataset_id != dataset.dataset_id:
        raise HydrocarbonInterpretationExportError("Отчёт относится к другому набору данных")
    expected_rows = dataset.active_index.values.size
    invalid_curves = tuple(
        curve.metadata.original_mnemonic
        for curve in dataset.curves.values()
        if curve.values.size != expected_rows
    )
    if invalid_curves:
        names = ", ".join(invalid_curves[:5])
        suffix = "…" if len(invalid_curves) > 5 else ""
        raise HydrocarbonInterpretationExportError(
            f"Кривые с неверным числом отсчётов: {names}{suffix}"
        )


def _prepare_target(target: str | Path, suffix: str, *, overwrite: bool) -> Path:
    destination = Path(target)
    if destination.suffix.casefold() != suffix:
        destination = destination.with_suffix(suffix)
    if destination.exists() and not overwrite:
        raise FileExistsError(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def _temporary_path(destination: Path) -> Path:
    descriptor, name = tempfile.mkstemp(
        prefix=f".{destination.stem}-",
        suffix=destination.suffix,
        dir=destination.parent,
    )
    os.close(descriptor)
    return Path(name)
