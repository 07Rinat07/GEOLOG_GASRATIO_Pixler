from __future__ import annotations

import csv
import math
from collections.abc import Mapping
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]

from geoworkbench.calculations.interval_statistics import CurveIntervalStatistics
from geoworkbench.data.spreadsheet_safety import protect_spreadsheet_row
from geoworkbench.services.localization import AppLanguage, Localizer


def statistics_columns(language: AppLanguage = AppLanguage.EN) -> tuple[str, ...]:
    localizer = Localizer.create(language)
    return (
        localizer.text("statistics.parameter"),
        localizer.text("statistics.mnemonic"),
        localizer.text("statistics.unit"),
        localizer.text("statistics.availability"),
        localizer.text("statistics.points"),
        localizer.text("statistics.zeros"),
        localizer.text("statistics.missing"),
        localizer.text("statistics.coverage"),
        localizer.text("statistics.minimum"),
        localizer.text("statistics.maximum"),
        localizer.text("statistics.mean"),
    )


def statistics_rows(
    statistics: tuple[CurveIntervalStatistics, ...],
    *,
    display_names: Mapping[str, str] | None = None,
    language: AppLanguage = AppLanguage.EN,
) -> tuple[tuple[object, ...], ...]:
    labels = display_names or {}
    localizer = Localizer.create(language)
    return tuple(
        (
            labels.get(item.mnemonic, item.mnemonic),
            item.mnemonic,
            item.unit or "",
            localizer.text(
                "statistics.available"
                if item.availability.value == "available"
                else "statistics.unavailable"
            ),
            item.valid_count,
            item.zero_count,
            item.missing_count if item.missing_count is not None else max(0, (item.total_count or item.valid_count) - item.valid_count),
            item.coverage_percent,
            _finite_or_none(item.minimum),
            _finite_or_none(item.maximum),
            _finite_or_none(item.mean),
        )
        for item in statistics
    )


def statistics_tsv(
    statistics: tuple[CurveIntervalStatistics, ...],
    *,
    interval_label: str,
    dataset_name: str,
    display_names: Mapping[str, str] | None = None,
    language: AppLanguage = AppLanguage.EN,
) -> str:
    localizer = Localizer.create(language)
    lines = [
        _tsv_row((localizer.text("statistics.dataset_header"), dataset_name)),
        _tsv_row((localizer.text("statistics.interval_header"), interval_label)),
    ]
    lines.append(_tsv_row(statistics_columns(language)))
    for row in statistics_rows(statistics, display_names=display_names, language=language):
        lines.append(_tsv_row(row))
    return "\n".join(lines)


def export_interval_statistics_csv(
    path: str | Path,
    statistics: tuple[CurveIntervalStatistics, ...],
    *,
    interval_label: str,
    dataset_name: str,
    display_names: Mapping[str, str] | None = None,
    language: AppLanguage = AppLanguage.EN,
) -> Path:
    target = Path(path)
    localizer = Localizer.create(language)
    with target.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(
            protect_spreadsheet_row(
                (localizer.text("statistics.dataset_header"), dataset_name)
            )
        )
        writer.writerow(
            protect_spreadsheet_row(
                (localizer.text("statistics.interval_header"), interval_label)
            )
        )
        writer.writerow(())
        writer.writerow(protect_spreadsheet_row(statistics_columns(language)))
        writer.writerows(
            protect_spreadsheet_row(row)
            for row in statistics_rows(
                statistics,
                display_names=display_names,
                language=language,
            )
        )
    return target


def export_interval_statistics_xlsx(
    path: str | Path,
    statistics: tuple[CurveIntervalStatistics, ...],
    *,
    interval_label: str,
    dataset_name: str,
    display_names: Mapping[str, str] | None = None,
    language: AppLanguage = AppLanguage.EN,
) -> Path:
    target = Path(path)
    localizer = Localizer.create(language)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = localizer.text("statistics.sheet_title")
    sheet.append(
        protect_spreadsheet_row((localizer.text("statistics.dataset_header"), dataset_name))
    )
    sheet.append(
        protect_spreadsheet_row((localizer.text("statistics.interval_header"), interval_label))
    )
    sheet.append(())
    sheet.append(protect_spreadsheet_row(statistics_columns(language)))
    for row in statistics_rows(statistics, display_names=display_names, language=language):
        sheet.append(protect_spreadsheet_row(row))

    header_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in sheet[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = (34, 18, 14, 16, 14, 12, 14, 14, 16, 16, 16)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=5, min_col=8, max_col=11):
        for cell in row:
            cell.number_format = "0.########"
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:K{max(4, sheet.max_row)}"
    workbook.save(target)
    return target


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.12g}"
    return str(value)


def _tsv_row(values: tuple[object, ...]) -> str:
    return "\t".join(_text(value) for value in protect_spreadsheet_row(values))


def _finite_or_none(value: float) -> float | None:
    return value if math.isfinite(value) else None
