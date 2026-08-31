from __future__ import annotations

import os
from pathlib import Path
import tempfile
import zipfile

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from geoworkbench.data.hydrocarbon_interpretation_export import (
    _docx_styles,
    _paragraph,
    _table,
)
from geoworkbench.data.spreadsheet_safety import protect_spreadsheet_row
from geoworkbench.printing.interpretation_report import (
    LBA_FIELDS,
    AnalysisInterpretationEntry,
    GeologicalGasStatistics,
    GeologicalRockComponent,
    GeologicalStratigraphyEntry,
    InterpretationReport,
    _LABELS,
    _LBA_LABELS,
)
from geoworkbench.services.lba_standard import describe_lba_assessment
from geoworkbench.services.localization import AppLanguage
from geoworkbench.services.report_output_transaction import (
    execute_report_output_transaction,
)
from geoworkbench.services.report_passport import ReportPassport


class InterpretationReportOfficeError(RuntimeError):
    pass


def export_interpretation_report_xlsx(
    report: InterpretationReport,
    target: str | Path,
    *,
    language: AppLanguage = AppLanguage.RU,
    overwrite: bool = False,
    passport: ReportPassport | None = None,
) -> Path:
    destination = _normalized_target(target, ".xlsx")
    if passport is not None:
        result = execute_report_output_transaction(
            destination,
            lambda staged: export_interpretation_report_xlsx(
                report,
                staged,
                language=language,
                overwrite=True,
            ),
            passport,
            overwrite=overwrite,
        )
        return result.primary_path
    _ensure_target(destination, overwrite=overwrite)
    temporary = _temporary_path(destination)
    try:
        _write_xlsx(temporary, report, language)
        os.replace(temporary, destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        if isinstance(exc, (FileExistsError, InterpretationReportOfficeError)):
            raise
        raise InterpretationReportOfficeError(
            f"Не удалось экспортировать Excel: {destination}"
        ) from exc
    return destination


def export_interpretation_report_docx(
    report: InterpretationReport,
    target: str | Path,
    *,
    language: AppLanguage = AppLanguage.RU,
    overwrite: bool = False,
    passport: ReportPassport | None = None,
) -> Path:
    destination = _normalized_target(target, ".docx")
    if passport is not None:
        result = execute_report_output_transaction(
            destination,
            lambda staged: export_interpretation_report_docx(
                report,
                staged,
                language=language,
                overwrite=True,
            ),
            passport,
            overwrite=overwrite,
        )
        return result.primary_path
    _ensure_target(destination, overwrite=overwrite)
    temporary = _temporary_path(destination)
    try:
        _write_docx(temporary, report, language)
        os.replace(temporary, destination)
    except Exception as exc:
        temporary.unlink(missing_ok=True)
        if isinstance(exc, (FileExistsError, InterpretationReportOfficeError)):
            raise
        raise InterpretationReportOfficeError(
            f"Не удалось экспортировать Word: {destination}"
        ) from exc
    return destination


def _write_xlsx(
    path: Path, report: InterpretationReport, language: AppLanguage
) -> None:
    labels = _LABELS[language]
    workbook = Workbook()
    summary = workbook.active
    summary.title = _sheet_name(labels["summary"])
    _append_rows(
        summary,
        (
            (labels["title"], ""),
            (labels["project"], report.project_name),
            (labels["well"], report.well_name),
            (labels["dataset"], report.dataset_name or "—"),
            (labels["actual_samples"], report.sample_count),
            (labels["calcimetry"], report.calcimetry_count),
            (labels["lba"], report.lba_count),
            (labels["interpretation"], report.interpreted_count),
        ),
    )
    summary["A1"].font = Font(bold=True, size=16)
    summary.merge_cells("A1:B1")
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 80

    meter = workbook.create_sheet(_sheet_name(labels["meter_section"]))
    _write_table(
        meter,
        (
            labels["interval"],
            labels["sample_intervals"],
            labels["coverage"],
            labels["composition"],
            labels["rock_description"],
            labels["stratigraphy"],
        ),
        tuple(
            (
                _interval(entry.top_depth, entry.bottom_depth, report.depth_unit),
                "\n".join(
                    _interval(top, bottom, report.depth_unit)
                    for top, bottom in entry.sample_intervals
                ),
                entry.sampling_coverage,
                _components(entry.rock_components, language),
                "\n\n".join(entry.rock_descriptions),
                _stratigraphy(entry.stratigraphy, language, report.depth_unit),
            )
            for entry in report.meter_geology
        ),
        widths=(18, 25, 15, 34, 70, 42),
    )
    for cell in meter["C"][1:]:
        cell.number_format = "0.0%"

    samples = workbook.create_sheet(_sheet_name(labels["sample_section"]))
    lba_headers = tuple(_LBA_LABELS[language][key] for key, _ in LBA_FIELDS)
    sample_headers = (
        labels["interval"],
        labels["composition"],
        labels["rock_description"],
        labels["stratigraphy"],
        "CaCO3, %",
        "CaMg(CO3)2, %",
        f"{labels['insoluble']}, %",
        *lba_headers,
        labels["lba_standard"],
        labels["interpretation"],
    )
    _write_table(
        samples,
        sample_headers,
        tuple(_sample_row(entry, report, language) for entry in report.entries),
        widths=(18, 34, 60, 40, 13, 15, 18, *((20,) * len(lba_headers)), 45, 60),
    )

    gas = workbook.create_sheet(_sheet_name(labels["gas_lba_section"]))
    _write_table(
        gas,
        (
            labels["interval"],
            labels["name"],
            "Mnemonic",
            "Unit",
            labels["minimum"],
            labels["mean"],
            labels["maximum"],
        ),
        tuple(
            _gas_row(entry, statistic, report.depth_unit, labels)
            for entry in report.entries
            for statistic in entry.gas_statistics
        ),
        widths=(18, 28, 18, 12, 16, 16, 16),
    )

    stratigraphy = workbook.create_sheet(_sheet_name(labels["stratigraphy_section"]))
    _write_table(
        stratigraphy,
        (
            labels["interval"],
            labels["rank"],
            labels["code"],
            labels["name"],
            labels["description"],
        ),
        tuple(
            (
                _interval(item.top_depth, item.bottom_depth, report.depth_unit),
                item.rank or "",
                item.code,
                item.localized_name(language),
                item.description or "",
            )
            for item in report.stratigraphy
        ),
        widths=(18, 18, 16, 32, 70),
    )
    workbook.save(path)


def _sample_row(
    entry: AnalysisInterpretationEntry,
    report: InterpretationReport,
    language: AppLanguage,
) -> tuple[object, ...]:
    observations = dict(entry.lba_observations)
    assessment = (
        describe_lba_assessment(entry.lba_standard_assessment, language)
        if entry.lba_standard_assessment is not None
        else ""
    )
    return (
        _interval(entry.top_depth, entry.bottom_depth, report.depth_unit),
        _components(entry.rock_components, language),
        entry.rock_description or "",
        _stratigraphy(entry.stratigraphy, language, report.depth_unit),
        entry.calcite_percent,
        entry.dolomite_percent,
        entry.insoluble_residue_percent,
        *(observations.get(key, "") for key, _ in LBA_FIELDS),
        assessment,
        entry.interpretation or "",
    )


def _gas_row(
    entry: AnalysisInterpretationEntry,
    statistic: GeologicalGasStatistics,
    depth_unit: str,
    labels: dict[str, str],
) -> tuple[object, ...]:
    name = {
        "total": labels["gas_total"],
        "sum": labels["gas_component_sum"],
    }.get(statistic.kind, statistic.mnemonic)
    return (
        _interval(entry.top_depth, entry.bottom_depth, depth_unit),
        name,
        statistic.mnemonic,
        statistic.unit,
        statistic.minimum,
        statistic.mean,
        statistic.maximum,
    )


def _write_docx(
    path: Path, report: InterpretationReport, language: AppLanguage
) -> None:
    labels = _LABELS[language]
    body = [
        _paragraph(labels["title"], style="Title"),
        _paragraph(f"{labels['project']}: {report.project_name}"),
        _paragraph(f"{labels['well']}: {report.well_name}"),
        _paragraph(f"{labels['dataset']}: {report.dataset_name or '—'}"),
        _paragraph(labels["meter_section"], style="Heading1"),
        _table(
            (
                labels["interval"],
                labels["sample_intervals"],
                labels["coverage"],
                labels["composition"],
                labels["rock_description"],
                labels["stratigraphy"],
            ),
            tuple(
                (
                    _interval(item.top_depth, item.bottom_depth, report.depth_unit),
                    "; ".join(
                        _interval(top, bottom, report.depth_unit)
                        for top, bottom in item.sample_intervals
                    ),
                    f"{item.sampling_coverage * 100:g}%",
                    _components(item.rock_components, language),
                    " | ".join(item.rock_descriptions),
                    _stratigraphy(item.stratigraphy, language, report.depth_unit),
                )
                for item in report.meter_geology
            ),
            widths=(1500, 1800, 1100, 3000, 4200, 3500),
        ),
        _paragraph(labels["sample_section"], style="Heading1"),
        _table(
            (
                labels["interval"],
                labels["composition"],
                labels["rock_description"],
                labels["stratigraphy"],
                labels["calcimetry"],
                labels["lba"],
                labels["gas"],
                labels["interpretation"],
            ),
            tuple(_docx_sample_row(item, report, language) for item in report.entries),
            widths=(1100, 1800, 2400, 1800, 1300, 2200, 2500, 2000),
        ),
        _paragraph(labels["stratigraphy_section"], style="Heading1"),
        _table(
            (
                labels["interval"],
                labels["rank"],
                labels["code"],
                labels["name"],
                labels["description"],
            ),
            tuple(
                (
                    _interval(item.top_depth, item.bottom_depth, report.depth_unit),
                    item.rank or "",
                    item.code,
                    item.localized_name(language),
                    item.description or "",
                )
                for item in report.stratigraphy
            ),
            widths=(2200, 1800, 1600, 3600, 5900),
        ),
        _paragraph(labels["notice"]),
    ]
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + "".join(body)
        + '<w:sectPr><w:pgSz w:w="16838" w:h="11906" w:orient="landscape"/>'
        '<w:pgMar w:top="700" w:right="700" w:bottom="700" w:left="700" '
        'w:header="500" w:footer="500" w:gutter="0"/></w:sectPr>'
        "</w:body></w:document>"
    )
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            '<Override PartName="/word/styles.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
            "</Types>",
        )
        package.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="word/document.xml"/></Relationships>',
        )
        package.writestr("word/document.xml", document_xml)
        package.writestr(
            "word/_rels/document.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
        )
        package.writestr("word/styles.xml", _docx_styles())


def _docx_sample_row(
    entry: AnalysisInterpretationEntry,
    report: InterpretationReport,
    language: AppLanguage,
) -> tuple[str, ...]:
    labels = _LABELS[language]
    lba_labels = _LBA_LABELS[language]
    observations = dict(entry.lba_observations)
    lba = "; ".join(
        f"{lba_labels[key]}: {observations.get(key, '—')}"
        for key, _ in LBA_FIELDS
    )
    if entry.lba_standard_assessment is not None:
        lba += "; " + describe_lba_assessment(entry.lba_standard_assessment, language)
    calcimetry = "; ".join(
        f"{name}: {value:g}%"
        for name, value in (
            ("CaCO3", entry.calcite_percent),
            ("CaMg(CO3)2", entry.dolomite_percent),
            (labels["insoluble"], entry.insoluble_residue_percent),
        )
        if value is not None
    )
    gas = "; ".join(
        _gas_text(item, labels) for item in entry.gas_statistics
    )
    return (
        _interval(entry.top_depth, entry.bottom_depth, report.depth_unit),
        _components(entry.rock_components, language),
        entry.rock_description or "",
        _stratigraphy(entry.stratigraphy, language, report.depth_unit),
        calcimetry,
        lba,
        gas,
        entry.interpretation or "",
    )


def _gas_text(item: GeologicalGasStatistics, labels: dict[str, str]) -> str:
    name = {
        "total": labels["gas_total"],
        "sum": labels["gas_component_sum"],
    }.get(item.kind, item.mnemonic)
    unit = f" {item.unit}" if item.unit else ""
    return (
        f"{name}: {labels['minimum']} {_number(item.minimum)}; "
        f"{labels['mean']} {_number(item.mean)}; "
        f"{labels['maximum']} {_number(item.maximum)}{unit}"
    )


def _components(
    components: tuple[GeologicalRockComponent, ...], language: AppLanguage
) -> str:
    return "\n".join(
        f"{item.localized_name(language)} ({item.code}): {item.percentage:g}%"
        for item in components
    )


def _stratigraphy(
    entries: tuple[GeologicalStratigraphyEntry, ...],
    language: AppLanguage,
    depth_unit: str,
) -> str:
    return "\n".join(
        f"{item.rank or '—'}; {item.code} — {item.localized_name(language)} "
        f"({_interval(item.top_depth, item.bottom_depth, depth_unit)})"
        for item in entries
    )


def _interval(top: float, bottom: float, unit: str) -> str:
    return f"{top:g}–{bottom:g}{f' {unit}' if unit else ''}"


def _number(value: float | None) -> str:
    return "—" if value is None else f"{value:.6g}"


def _append_rows(worksheet, rows: tuple[tuple[object, ...], ...]) -> None:
    for row in rows:
        worksheet.append(protect_spreadsheet_row(row))


def _write_table(
    worksheet,
    headers: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
    *,
    widths: tuple[int, ...],
) -> None:
    worksheet.append(protect_spreadsheet_row(headers))
    for row in rows:
        worksheet.append(protect_spreadsheet_row(row))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="172033")
        cell.fill = PatternFill("solid", fgColor="DCE8F4")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _sheet_name(value: str) -> str:
    sanitized = "".join("_" if char in r"[]:*?/\\" else char for char in value)
    return (sanitized.strip() or "Report")[:31]


def _normalized_target(target: str | Path, suffix: str) -> Path:
    destination = Path(target)
    return destination if destination.suffix.casefold() == suffix else destination.with_suffix(suffix)


def _ensure_target(destination: Path, *, overwrite: bool) -> None:
    if destination.exists() and not overwrite:
        raise FileExistsError(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)


def _temporary_path(destination: Path) -> Path:
    descriptor, name = tempfile.mkstemp(
        prefix=f".{destination.stem}-",
        suffix=destination.suffix,
        dir=destination.parent,
    )
    os.close(descriptor)
    return Path(name)


__all__ = [
    "InterpretationReportOfficeError",
    "export_interpretation_report_docx",
    "export_interpretation_report_xlsx",
]
